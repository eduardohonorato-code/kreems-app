"""
Componentes reutilizables para todas las páginas.
"""
import streamlit as st
import base64
import os
from datetime import date

FONT_FAMILY = "Inter, sans-serif"

# ── FUENTE GLOBAL ─────────────────────────────────────────────

def inject_font():
    """Inyecta Inter desde Google Fonts como fuente global de la app."""
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
        @import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Rounded:opsz,wght,FILL,GRAD@20..48,100..700,0..1,-50..200&display=block');

        html, body, [class*="css"], [class*="st-"],
        .stMarkdown, .stText, .stDataFrame,
        button, input, select, textarea,
        h1, h2, h3, h4, h5, h6, p, span, div, label {
            font-family: 'Inter', sans-serif !important;
            font-variant-ligatures: none !important;
            font-feature-settings: "liga" 0, "clig" 0 !important;
        }

        /* ── Menú de columnas del dataframe (cobertura total) ── */

        /* Contenedor raíz del popup */
        .ag-popup { position: absolute !important; z-index: 999999 !important; }
        .ag-popup > * { background-color: #FFFFFF !important; }

        /* Menú principal y todas sus capas */
        .ag-menu,
        .ag-column-menu,
        .ag-menu-body,
        .ag-menu-list,
        .ag-tabs,
        .ag-tabs-header,
        .ag-tabs-body,
        .ag-tab-body,
        .ag-menu-column-select,
        .ag-menu-column-select-wrapper {
            background-color: #FFFFFF !important;
            color: #0F172A !important;
        }

        /* Borde y sombra solo al contenedor más externo del menú */
        .ag-menu {
            border: 1px solid #E2E8F0 !important;
            border-radius: 8px !important;
            box-shadow: 0 8px 24px rgba(0,0,0,0.12) !important;
            overflow: hidden !important;
            z-index: 999999 !important;
        }

        /* Ítems del menú */
        .ag-menu-option,
        .ag-menu-option-part {
            background-color: #FFFFFF !important;
            color: #0F172A !important;
            font-size: 13px !important;
        }
        .ag-menu-option:hover { background-color: #F1F5F9 !important; }
        .ag-menu-option--active { background-color: #F1F5F9 !important; }
        .ag-menu-option-icon { color: #94A3B8 !important; background: transparent !important; }
        .ag-menu-separator { border-color: #E2E8F0 !important; }

        /* Tabs (Filtrar / Columnas) */
        .ag-tab {
            background-color: #FFFFFF !important;
            color: #94A3B8 !important;
            border-bottom: 2px solid transparent !important;
        }
        .ag-tab--selected, .ag-tab-selected {
            color: #c4007a !important;
            border-bottom-color: #c4007a !important;
            background-color: #FFFFFF !important;
        }

        /* Input de filtro */
        .ag-filter,
        .ag-filter-body,
        .ag-filter-body-wrapper { background-color: #FFFFFF !important; }
        .ag-filter input, .ag-floating-filter-input {
            background-color: #F7F8FA !important;
            border: 1px solid #E2E8F0 !important;
            border-radius: 6px !important;
            color: #0F172A !important;
        }

        /* Forzar overflow visible en el stDataFrame para que el popup no quede cortado */
        [data-testid="stDataFrame"] > div { overflow: visible !important; }

        /* ── Sidebar oscuro aplicado directamente por CSS ── */
        [data-testid="stSidebar"] {
            background-color: #18112E !important;
            border-right: none !important;
            box-shadow: 2px 0 12px rgba(0,0,0,0.20) !important;
        }
        [data-testid="stSidebar"] > div:first-child {
            background-color: #18112E !important;
        }
        [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] span,
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] div {
            color: #C8BEE0 !important;
        }

        /* Links de navegación */
        [data-testid="stSidebar"] [data-testid="stPageLink"] a,
        [data-testid="stSidebar"] [data-testid="stPageLink"] p {
            color: #C8BEE0 !important;
            font-weight: 500 !important;
            font-size: 13.5px !important;
            transition: color 0.15s ease;
        }
        [data-testid="stSidebar"] [data-testid="stPageLink"]:hover a,
        [data-testid="stSidebar"] [data-testid="stPageLink"]:hover p {
            color: #FFFFFF !important;
        }

        /* Radio buttons */
        [data-testid="stSidebar"] .stRadio label,
        [data-testid="stSidebar"] .stRadio p {
            color: #C8BEE0 !important;
            font-size: 13px !important;
        }

        /* Botón cerrar sesión */
        [data-testid="stSidebar"] .stButton button {
            background: rgba(255,255,255,0.06) !important;
            border: 1px solid rgba(255,255,255,0.14) !important;
            color: #C8BEE0 !important;
            font-size: 13px !important;
        }
        [data-testid="stSidebar"] .stButton button:hover {
            background: rgba(255,255,255,0.12) !important;
            color: #FFFFFF !important;
        }

        /* Checkboxes */
        [data-testid="stSidebar"] .stCheckbox label,
        [data-testid="stSidebar"] .stCheckbox p {
            color: #C8BEE0 !important;
        }

        /* ── Botón colapsar sidebar (dentro del sidebar) ── */
        [data-testid="stSidebarCollapseButton"] {
            width: 2rem !important; height: 2rem !important;
            overflow: hidden !important; position: relative !important;
            border-radius: 6px !important;
        }
        [data-testid="stSidebarCollapseButton"] span {
            font-size: 0 !important; opacity: 0 !important;
            position: absolute !important; pointer-events: none !important;
        }
        [data-testid="stSidebarCollapseButton"]::before {
            content: "☰";
            font-size: 16px !important; font-family: Arial, sans-serif !important;
            color: rgba(200,190,224,0.75) !important;
            position: absolute !important;
            top: 50% !important; left: 50% !important;
            transform: translate(-50%, -50%) !important;
        }
        [data-testid="stSidebarCollapseButton"]:hover::before { color: #FFFFFF !important; }

        /* ── Botón expandir sidebar (fuera del sidebar, cuando está colapsado) ── */
        [data-testid="stExpandSidebarButton"] {
            width: 2rem !important; height: 2rem !important;
            overflow: hidden !important; position: relative !important;
            border-radius: 6px !important;
        }
        [data-testid="stExpandSidebarButton"] span {
            font-size: 0 !important; opacity: 0 !important;
            position: absolute !important; pointer-events: none !important;
        }
        [data-testid="stExpandSidebarButton"]::before {
            content: "☰";
            font-size: 16px !important; font-family: Arial, sans-serif !important;
            color: #2d0050 !important;
            position: absolute !important;
            top: 50% !important; left: 50% !important;
            transform: translate(-50%, -50%) !important;
        }
        [data-testid="stExpandSidebarButton"]:hover::before { color: #c4007a !important; }
    </style>
    """, unsafe_allow_html=True)


# ── LOGO ──────────────────────────────────────────────────────

@st.cache_data
def _logo_b64() -> str | None:
    """Carga el logo como base64. Retorna None si no existe el archivo."""
    logo_path = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "assets", "logo.png")
    )
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    return None


def _logo_html(altura: int = 40) -> str:
    """Retorna el HTML del logo: imagen si existe, texto de fallback si no."""
    b64 = _logo_b64()
    if b64:
        return f"<img src='data:image/png;base64,{b64}' style='height:{altura}px;'/>"
    # Fallback: texto estilizado
    return "<span style='color:#c4007a; font-size:2rem; font-weight:800; letter-spacing:-1px;'>kreems</span>"

MESES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}

# ── AÑO FISCAL — cambiar aquí para el siguiente año ──────────
_HOY       = date.today()
ANO_FISCAL = _HOY.year          # detecta el año actual automáticamente
MES_NUM_ACTUAL = _HOY.month

MESES_PERIODOS = {
    nombre: f"{ANO_FISCAL}-{num:02d}"
    for num, nombre in {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }.items()
}

# ── CONSTANTES DE NEGOCIO (fuente única de verdad) ────────────
# Nombres de sociedad: usar SIEMPRE estas constantes, nunca literales sueltos.
# Evita inconsistencias como 'ACUNA' vs 'ACUÑA' que rompen los filtros SQL.
SOC_ACUNA        = "ACUÑA"
SOC_GRAN_NATURAL = "GRAN_NATURAL"
SOCIEDADES       = [SOC_GRAN_NATURAL, SOC_ACUNA]
OPCIONES_SOCIEDAD = ["Todas"] + SOCIEDADES

# Mapa código CC → nombre. Único lugar donde se define.
NOMBRES_CC = {
    "CC-01": "Administración",
    "CC-02": "Comercial",
    "CC-03": "Distribución",
    "CC-04": "Producción",
}


def get_soc_sql_filter(sociedad_sel: str) -> str:
    """Fragmento SQL para filtrar por sociedad.
    Retorna '' si es 'Todas'. Los valores provienen de OPCIONES_SOCIEDAD
    (lista cerrada), por lo que no hay entrada de texto libre del usuario.
    """
    return f"AND sociedad = '{sociedad_sel}'" if sociedad_sel != "Todas" else ""


def _estado_mes(num: int) -> tuple[str, str, str]:
    """Retorna (label, color, icono) según el estado del mes."""
    if num < MES_NUM_ACTUAL:
        return ("CERRADO", "#0F6E56", "✓")
    elif num == MES_NUM_ACTUAL:
        return ("ACTUAL", "#c4007a", "●")
    else:
        return ("PROYEC.", "#aaa", "◌")


# ── HEADER ────────────────────────────────────────────────────

def header(titulo: str):
    """Header con logo Kreems y título de página."""
    inject_font()
    st.markdown(f"""
    <div style="
        display: flex;
        align-items: center;
        justify-content: space-between;
        background: #FFFFFF;
        border-bottom: 1px solid #E2E8F0;
        padding: 12px 24px;
        margin: -1rem -1rem 1.5rem -1rem;
    ">
        {_logo_html(altura=42)}
        <span style="color:#2d0050; font-size:1.3rem; font-weight:600;">{titulo}</span>
        <span style="color:#bbb; font-size:12px;">Control Presupuestario {ANO_FISCAL}</span>
    </div>
    """, unsafe_allow_html=True)


# ── SELECTOR MESES ────────────────────────────────────────────

def selector_meses(key: str = "mes", default: str = "Mayo") -> tuple:
    """
    Selector de período con botones rápidos por mes + rango personalizado Desde/Hasta.
    Retorna (periodo_desde, periodo_hasta) en formato 'YYYY-MM'.
    """
    meses_lista   = list(MESES.values())
    mes_actual_nm = MESES[MES_NUM_ACTUAL]

    k_desde = f"_sm_{key}_desde"
    k_hasta = f"_sm_{key}_hasta"
    if k_desde not in st.session_state:
        st.session_state[k_desde] = default
    if k_hasta not in st.session_state:
        st.session_state[k_hasta] = default

    desde_act = st.session_state[k_desde]
    hasta_act  = st.session_state[k_hasta]

    # ── Fila 1: botones rápidos por mes (solo hasta el mes actual) + YTD ──
    meses_visibles = [(num, nom) for num, nom in MESES.items() if num <= MES_NUM_ACTUAL]
    n = len(meses_visibles)
    # columnas: n meses + separador + YTD + label período
    cols_row1 = st.columns([1] * n + [0.25, 0.9, 2.2])

    for i, (num, nombre) in enumerate(meses_visibles):
        with cols_row1[i]:
            lbl, color, icono = _estado_mes(num)
            st.markdown(
                f"<div style='font-size:8px;color:{color};text-align:center;"
                f"font-weight:700;letter-spacing:0.3px;padding-bottom:2px;'>"
                f"{icono} {lbl}</div>", unsafe_allow_html=True
            )
            activo = (desde_act == nombre and hasta_act == nombre)
            if st.button(nombre[:3], key=f"_sm_{key}_m{num}",
                         type="primary" if activo else "secondary",
                         use_container_width=True):
                st.session_state[k_desde] = nombre
                st.session_state[k_hasta] = nombre
                st.rerun()

    # Separador visual
    with cols_row1[n]:
        st.markdown("<div style='padding-top:28px;text-align:center;color:#E2E8F0;font-size:18px;'>│</div>",
                    unsafe_allow_html=True)

    # Botón YTD
    with cols_row1[n + 1]:
        st.markdown("<div style='font-size:8px;color:#94A3B8;text-align:center;"
                    "font-weight:700;padding-bottom:2px;'>ACUM.</div>", unsafe_allow_html=True)
        ytd_activo = (desde_act == "Enero" and hasta_act == mes_actual_nm)
        if st.button("YTD", key=f"_sm_{key}_ytd",
                     type="primary" if ytd_activo else "secondary",
                     use_container_width=True):
            st.session_state[k_desde] = "Enero"
            st.session_state[k_hasta] = mes_actual_nm
            st.rerun()

    # Espacio reservado para el label (se rellena después con st.empty)
    with cols_row1[n + 2]:
        label_placeholder = st.empty()

    # ── Fila 2: rango personalizado Desde / Hasta ──────────────
    col_lbl, col_desde, col_arrow, col_hasta, col_spacer = \
        st.columns([0.6, 1.3, 0.12, 1.3, 3.0])

    with col_lbl:
        st.markdown("<div style='padding-top:7px;font-size:12px;color:#94A3B8;'>Rango</div>",
                    unsafe_allow_html=True)

    with col_desde:
        desde = st.selectbox("Desde", meses_lista,
            index=meses_lista.index(st.session_state[k_desde]),
            label_visibility="collapsed")
        st.session_state[k_desde] = desde

    with col_arrow:
        st.markdown("<div style='padding-top:8px;text-align:center;color:#CBD5E1;'>→</div>",
                    unsafe_allow_html=True)

    with col_hasta:
        hasta = st.selectbox("Hasta", meses_lista,
            index=meses_lista.index(st.session_state[k_hasta]),
            label_visibility="collapsed")
        st.session_state[k_hasta] = hasta

    # Validar rango
    if meses_lista.index(hasta) < meses_lista.index(desde):
        hasta = desde
        st.session_state[k_hasta] = desde

    # Label con valores finales correctos (después de leer los selectboxes)
    periodo_txt = desde if desde == hasta else f"{desde} – {hasta}"
    label_placeholder.markdown(
        f"<div style='padding-top:26px;font-size:13px;color:#c4007a;"
        f"font-weight:600;'>📅 {periodo_txt} {ANO_FISCAL}</div>",
        unsafe_allow_html=True
    )

    return (MESES_PERIODOS[desde], MESES_PERIODOS[hasta])


# ── KPI CARD ──────────────────────────────────────────────────

def kpi_card(label: str, valor: float, referencia: float = None,
             prefijo: str = "$", sufijo: str = "", decimales: int = 0,
             invertir_color: bool = False):
    """Tarjeta KPI con valor, referencia y variación %."""
    fmt = f"{{:,.{decimales}f}}"
    val_fmt = prefijo + fmt.format(abs(valor) / 1_000_000) + "M" + sufijo

    if referencia and referencia != 0:
        variacion = ((valor - referencia) / abs(referencia)) * 100
        ref_fmt = prefijo + fmt.format(abs(referencia) / 1_000_000) + "M"

        if invertir_color:
            color = "#0F6E56" if variacion <= 0 else "#c4007a"
            icono = "↓" if variacion <= 0 else "↑"
        else:
            color = "#0F6E56" if variacion >= 0 else "#c4007a"
            icono = "↑" if variacion >= 0 else "↓"

        sub_html = f"""
        <div style="font-size:12px; color:{color}; margin-top:4px;">
            {icono} {abs(variacion):.1f}%
        </div>
        <div style="font-size:11px; color:#aaa; margin-top:2px;">Obj: {ref_fmt}</div>
        """
    else:
        sub_html = ""

    st.markdown(f"""
    <div style="background:#FFFFFF; border:1px solid #E2E8F0; border-radius:12px;
                padding:16px 18px; text-align:center; height:100%;
                box-shadow:0 1px 4px rgba(0,0,0,0.06);">
        <div style="font-size:11px; color:#94A3B8; margin-bottom:6px; text-transform:uppercase; letter-spacing:0.5px;">{label}</div>
        <div style="font-size:22px; font-weight:700; color:#0F172A;">{val_fmt}</div>
        {sub_html}
    </div>
    """, unsafe_allow_html=True)


# ── SIDEBAR ───────────────────────────────────────────────────

def sidebar_kreems(mostrar_sociedad: bool = True, mostrar_cc: bool = False):
    """Sidebar estándar con logo, navegación y filtros."""
    with st.sidebar:
        # Logo
        st.markdown(f"""
        <div style="text-align:center; padding:20px 0 14px;">
            {_logo_html(altura=44)}
            <div style="color:rgba(255,255,255,0.38); font-size:11px; margin-top:8px; letter-spacing:0.3px;">
                Control Presupuestario {ANO_FISCAL}
            </div>
        </div>
        """, unsafe_allow_html=True)

        def _nav_section(label: str, top_pad: bool = False):
            pad = "10px" if top_pad else "0"
            st.markdown(f"""
            <div style="font-size:9px; color:rgba(255,255,255,0.3); font-weight:700;
                        letter-spacing:1.5px; padding:{pad} 4px 4px;
                        text-transform:uppercase;">{label}</div>
            """, unsafe_allow_html=True)

        # ── REPORTES ─────────────────────────────────────────────
        _nav_section("Reportes")
        st.page_link("app.py",                      label="🏠  Inicio",                   use_container_width=True)
        st.page_link("pages/1_dashboard.py",        label="📊  Resumen Ejecutivo",         use_container_width=True)
        st.page_link("pages/2_eerr.py",             label="📋  Estado de Resultados",      use_container_width=True)
        st.page_link("pages/12_eerr_acumulado.py",  label="📈  EERR Acumulado",            use_container_width=True)
        st.page_link("pages/13_eerr_mensual.py",    label="📅  EERR Mensual & Análisis",   use_container_width=True)
        st.page_link("pages/3_centro_costos.py",    label="🏢  Centro de Costos",          use_container_width=True)
        st.page_link("pages/4_control_cuentas.py",  label="📑  Control Cuentas",           use_container_width=True)

        # ── ANÁLISIS ─────────────────────────────────────────────
        _nav_section("Análisis", top_pad=True)
        st.page_link("pages/8_proyeccion_cierre.py", label="📈  Proyección al Cierre",     use_container_width=True)
        st.page_link("pages/10_riesgos.py",          label="⚠️  Riesgos y Oportunidades", use_container_width=True)
        st.page_link("pages/11_flash_report.py",     label="📄  Flash Report Mensual",     use_container_width=True)
        st.page_link("pages/6_reportes.py",          label="🤖  Reportes IA Guardados",    use_container_width=True)

        # ── ADMINISTRACIÓN (solo admin) ───────────────────────────
        rol = st.session_state.get("rol", "")
        if rol == "admin":
            _nav_section("Administración", top_pad=True)
            st.page_link("pages/5_cargar_datos.py",        label="⬆️  Cargar Datos",        use_container_width=True)
            st.page_link("pages/14_presupuesto_detalle.py", label="🧾  Presupuesto Detalle",  use_container_width=True)
            st.page_link("pages/15_presupuesto_ventas_cv.py", label="📈  Ppto Ventas & CV",    use_container_width=True)
            st.page_link("pages/7_admin_usuarios.py",      label="👤  Admin Usuarios",       use_container_width=True)

        st.markdown("<hr style='border:none; border-top:1px solid rgba(255,255,255,0.08); margin:14px 0;'>",
                    unsafe_allow_html=True)

        sociedad_sel = "Todas"
        cc_sel = []

        if mostrar_sociedad:
            st.markdown("**🏭 Sociedad**")
            sociedad_sel = st.radio(
                "sociedad",
                OPCIONES_SOCIEDAD,
                label_visibility="collapsed"
            )

        if mostrar_cc:
            st.markdown("**🏢 Centro de Costo**")
            for codigo, nombre_cc in NOMBRES_CC.items():
                if st.checkbox(nombre_cc, value=True, key=f"cc_{codigo}"):
                    cc_sel.append(codigo)

        st.markdown("<hr style='border:none; border-top:1px solid rgba(255,255,255,0.08); margin:14px 0;'>",
                    unsafe_allow_html=True)

        # Usuario
        nombre = st.session_state.get("nombre", "")
        rol_display = st.session_state.get("rol", "").capitalize()
        st.markdown(f"""
        <div style="background:rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.1); border-radius:8px; padding:10px 12px; margin-bottom:10px;">
            <div style="font-size:12px; font-weight:600; color:#F0EBF8;">👤 {nombre}</div>
            <div style="font-size:11px; color:rgba(255,255,255,0.38);">{rol_display}</div>
        </div>
        """, unsafe_allow_html=True)

        if st.button("Cerrar sesión", use_container_width=True):
            for k in ["usuario", "rol", "nombre", "cc_permitidos"]:
                st.session_state.pop(k, None)
            st.rerun()

    return sociedad_sel, cc_sel


# ── SEMÁFORO PRESUPUESTARIO ───────────────────────────────────
# Umbrales para CC y cuentas de gasto (menos ejecución = mejor)
_SEM_ROJO    = 100.0   # > 100% → sobre presupuesto
_SEM_AMARILLO = 90.0   # 90–100% → zona de alerta

def semaforo_texto(pct: float) -> str:
    """Retorna emoji + etiqueta para columna Estado en tablas."""
    if pct > _SEM_ROJO:
        return "🔴 Alerta"
    if pct >= _SEM_AMARILLO:
        return "🟡 Atención"
    return "🟢 OK"

def semaforo_badge(pct: float) -> str:
    """Badge HTML con semáforo de ejecución. Ideal para secciones HTML."""
    if pct > _SEM_ROJO:
        bg, color, txt = "rgba(204,0,0,0.10)", "#cc0000", "🔴 Alerta"
    elif pct >= _SEM_AMARILLO:
        bg, color, txt = "rgba(217,119,6,0.11)", "#d97706", "🟡 Atención"
    else:
        bg, color, txt = "rgba(15,110,86,0.10)", "#0F6E56", "🟢 OK"
    return (
        f"<span style='background:{bg}; color:{color}; border-radius:20px; "
        f"padding:2px 9px; font-size:11px; font-weight:700; white-space:nowrap;'>"
        f"{txt}</span>"
    )

def semaforo_style_cell(pct: float) -> str:
    """CSS inline para colorear celdas de % ejecución en Pandas Styler."""
    if pct > _SEM_ROJO:
        return "color:#cc0000; font-weight:700"
    if pct >= _SEM_AMARILLO:
        return "color:#d97706; font-weight:600"
    return "color:#0F6E56; font-weight:500"


# ── HELPERS VISUALES ──────────────────────────────────────────

def badge_html(texto: str, tipo: str = "normal") -> str:
    """
    Retorna HTML de un badge de estado.
    tipos: 'ok'=verde, 'warn'=rosa/rojo, 'neutral'=gris, 'normal'=morado
    """
    estilos = {
        "ok":      ("rgba(15,110,86,0.12)",  "#0F6E56"),
        "warn":    ("rgba(196,0,122,0.12)",  "#c4007a"),
        "neutral": ("#f0f0f0",              "#888"),
        "normal":  ("#f0e6f8",              "#7b2d8b"),
    }
    bg, txt = estilos.get(tipo, estilos["normal"])
    return (
        f"<span style='background:{bg}; color:{txt}; border-radius:20px; "
        f"padding:2px 10px; font-size:11px; font-weight:600; "
        f"white-space:nowrap;'>{texto}</span>"
    )


def progress_bar_html(pct: float, max_pct: float = 150) -> str:
    """Barra de progreso + porcentaje para usar en celdas."""
    clamp = min(abs(pct), max_pct)
    color = "#0F6E56" if pct <= 100 else "#c4007a"
    width = (clamp / max_pct) * 100
    return (
        f"<div style='background:#f5eef8; border-radius:4px; "
        f"height:7px; width:100%; margin-bottom:2px;'>"
        f"<div style='background:{color}; width:{width:.1f}%; "
        f"height:7px; border-radius:4px;'></div></div>"
        f"<div style='font-size:11px; color:{color}; font-weight:600;'>"
        f"{pct:.1f}%</div>"
    )


def cc_card(nombre: str, codigo: str, real: float, ppto: float) -> str:
    """Card HTML para un Centro de Costo con semáforo de 3 estados."""
    pct       = (real / ppto * 100) if ppto else 0
    variacion = real - ppto

    COLOR_GOOD = "#0F6E56"
    COLOR_WARN = "#d97706"
    COLOR_BAD  = "#cc0000"

    # ── semáforo 3 estados ──
    if pct > _SEM_ROJO:
        color_pct  = COLOR_BAD
        badge_bg   = "rgba(204,0,0,0.10)"
        bar_color  = COLOR_BAD
        border_col = "#ffd0d0"
        sem_label  = "🔴 Alerta"
        status_note = (
            f"<div style='font-size:10px; color:{COLOR_BAD}; margin-bottom:6px; font-weight:600;'>"
            f"Sobre presupuesto en {pct-100:.1f}%</div>"
        )
    elif pct >= _SEM_AMARILLO:
        color_pct  = COLOR_WARN
        badge_bg   = "rgba(217,119,6,0.11)"
        bar_color  = COLOR_WARN
        border_col = "#ffe9c5"
        sem_label  = "🟡 Atención"
        status_note = (
            f"<div style='font-size:10px; color:{COLOR_WARN}; margin-bottom:6px; font-weight:600;'>"
            f"Zona de alerta — {pct:.1f}% ejecutado</div>"
        )
    else:
        color_pct  = COLOR_GOOD
        badge_bg   = "rgba(15,110,86,0.10)"
        bar_color  = "#c4007a"
        border_col = "#f0dff0"
        sem_label  = "🟢 OK"
        status_note = "<div style='margin-bottom:12px;'></div>"

    color_var = COLOR_BAD if variacion > 0 else COLOR_GOOD
    icono_var = "&#9650;" if variacion > 0 else "&#9660;"
    bar_fill  = min(pct, 100)

    return (
        f"<div style='background:#fff; border:1px solid {border_col}; border-radius:14px;"
        f"padding:18px 20px;'>"
        # fila título + badges
        f"<div style='display:flex; justify-content:space-between; align-items:flex-start;"
        f"margin-bottom:10px;'>"
        f"<div>"
        f"<div style='font-size:11px; color:#aaa; font-weight:600; text-transform:uppercase;"
        f"letter-spacing:0.5px;'>{codigo}</div>"
        f"<div style='font-size:15px; font-weight:700; color:#2d0050; margin-top:2px;'>{nombre}</div>"
        f"</div>"
        f"<div style='display:flex; flex-direction:column; align-items:flex-end; gap:4px;'>"
        f"<span style='background:{badge_bg}; color:{color_pct}; border-radius:20px;"
        f"padding:2px 9px; font-size:12px; font-weight:700;'>{pct:.1f}%</span>"
        f"<span style='font-size:11px; font-weight:700;'>{sem_label}</span>"
        f"</div>"
        f"</div>"
        # barra de progreso
        f"<div style='background:#f0f0f0; border-radius:6px; height:7px; margin-bottom:6px;'>"
        f"<div style='background:{bar_color}; width:{bar_fill:.1f}%; height:7px; border-radius:6px;'>"
        f"</div></div>"
        f"{status_note}"
        # montos
        f"<div style='display:flex; justify-content:space-between;'>"
        f"<div><div style='font-size:10px; color:#bbb;'>Ejecutado</div>"
        f"<div style='font-size:14px; font-weight:700; color:#2d0050;'>${real/1_000_000:,.2f}M</div></div>"
        f"<div style='text-align:right;'><div style='font-size:10px; color:#bbb;'>Presupuesto</div>"
        f"<div style='font-size:14px; font-weight:600; color:#888;'>${ppto/1_000_000:,.2f}M</div></div>"
        f"</div>"
        f"<div style='margin-top:8px; font-size:11px; color:{color_var}; font-weight:600;'>"
        f"{icono_var} ${abs(variacion)/1_000_000:,.2f}M vs ppto</div>"
        f"</div>"
    )


# ── FORMATTERS ────────────────────────────────────────────────

def fmt_mill(v: float) -> str:
    """Formatea como millones con 2 decimales."""
    if v is None:
        return "—"
    return f"${v/1_000_000:,.2f}M"


def fmt_clp(v: float) -> str:
    """Formatea como CLP con separador de miles."""
    if v is None:
        return "—"
    return f"${v:,.0f}"


def df_to_excel_bytes(frames: dict) -> bytes:
    """
    Convierte uno o varios DataFrames a un archivo Excel en memoria.
    Args:
        frames: dict donde la clave es el nombre de la hoja y el valor es un DataFrame.
    Returns:
        bytes del archivo .xlsx
    """
    import io
    import pandas as pd
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, df in frames.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    return buf.getvalue()


def boton_excel(frames: dict, nombre_archivo: str, label: str = "⬇ Exportar Excel"):
    """
    Renderiza un st.download_button para exportar DataFrames a Excel.
    Args:
        frames:         dict hoja→DataFrame
        nombre_archivo: nombre del archivo sin extensión
        label:          texto del botón
    """
    datos = df_to_excel_bytes(frames)
    st.download_button(
        label=label,
        data=datos,
        file_name=f"{nombre_archivo}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )


def df_to_excel_pct_bytes(frames: dict, id_col: str = "Concepto") -> bytes:
    """
    Exporta DataFrames cuyos valores están en puntos porcentuales (ej. 35.2 = 35.2%).
    Convierte esas columnas a fracción (0.352) y aplica el formato de celda '0.0%',
    de modo que en Excel aparecen directamente como porcentaje y, al reformatear,
    no se inflan. Todas las columnas excepto `id_col` se tratan como porcentaje.
    """
    import io
    import pandas as pd
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, df in frames.items():
            df2 = df.copy()
            pct_cols = [c for c in df2.columns if c != id_col]
            for c in pct_cols:
                df2[c] = pd.to_numeric(df2[c], errors="coerce") / 100.0

            hoja = sheet_name[:31]
            df2.to_excel(writer, sheet_name=hoja, index=False)
            ws = writer.sheets[hoja]

            pos = {name: i for i, name in enumerate(df2.columns)}  # 0-based
            for c in pct_cols:
                letra = get_column_letter(pos[c] + 1)  # openpyxl es 1-based
                for fila in range(2, len(df2) + 2):     # fila 1 = encabezado
                    ws[f"{letra}{fila}"].number_format = "0.0%"
    return buf.getvalue()


def boton_excel_pct(frames: dict, nombre_archivo: str,
                    label: str = "⬇ Exportar Excel", id_col: str = "Concepto"):
    """Igual que boton_excel pero las columnas numéricas se exportan como porcentaje real de Excel."""
    datos = df_to_excel_pct_bytes(frames, id_col=id_col)
    st.download_button(
        label=label,
        data=datos,
        file_name=f"{nombre_archivo}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )
