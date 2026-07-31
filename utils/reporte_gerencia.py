"""
Reporte de Gerencia — Real vs Presupuesto por centro de costo, con análisis de brechas.

Consolida en un solo entregable lo que hoy está repartido en EERR, Centro de Costos
y Control por Cuenta: P&L acumulado contra el presupuesto de los MISMOS meses,
puente de EBIT descompuesto, apertura por centro de costo y clasificación
automática de cada brecha (recurrente / puntual / desfase / no presupuestada).

Dos salidas desde la misma estructura de datos (`construir_reporte`):
  - `to_excel(rep)` → workbook corporativo, una hoja por bloque de análisis.
  - `to_html(rep)`  → página autocontenida con los insights, para leer sin Excel.

Toda la lógica de cálculo es pura (recibe DataFrames, no toca Streamlit) para
poder validarla contra la base sin levantar la app.
"""
from __future__ import annotations

import html as _html
import io
import math
from datetime import datetime

import pandas as pd

from utils.db import query
from utils.components import (
    NOMBRES_CC, SOC_ACUNA, SOC_GRAN_NATURAL, ETIQUETA_SOCIEDAD,
)

# Etiqueta corta de sociedad para las columnas angostas del detalle
_SOC_CORTA = {SOC_ACUNA: "AC", SOC_GRAN_NATURAL: "GN"}

# CC-00 no es un centro de costo: ahí cuelgan las ventas y el costo variable.
# Se le pone una etiqueta explícita para que no aparezca como "Ninguno".
CC_SIN_ASIGNAR = "CC-00"
ETIQUETA_CC_SIN_ASIGNAR = "Sin centro de costo"


def _label_cc(codigo: str, nombre_fallback: str = "") -> str:
    if codigo == CC_SIN_ASIGNAR:
        return ETIQUETA_CC_SIN_ASIGNAR
    return NOMBRES_CC.get(codigo, nombre_fallback or codigo)


def _plural(n: int, singular: str, plural: str) -> str:
    return f"{n} {singular if n == 1 else plural}"

# ── CONSTANTES ────────────────────────────────────────────────

ABREV_MES = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
}

# Líneas del P&L: (etiqueta, clasificacion | None si es subtotal, clave_subtotal)
LINEAS_PL = [
    ("Ventas",                  "INGRESO",        None),
    ("Costo de Venta",          "COSTO_VAR",      None),
    ("Utilidad Bruta",          None,             "UB"),
    ("Costo Fijo",              "COSTO_FIJO",     None),
    ("OPEX",                    "OPEX",           None),
    ("EBIT",                    None,             "EBIT"),
    ("Gastos Financieros",      "FINANCIERO",     None),
    ("Gastos No Operacionales", "NO_OPERACIONAL", None),
    ("Utilidad Neta",           None,             "UN"),
]

CLASIFS = ["INGRESO", "COSTO_VAR", "COSTO_FIJO", "OPEX", "FINANCIERO", "NO_OPERACIONAL"]

# Nombre legible de cada clasificación para las tablas de detalle
ETIQUETA_LINEA = {
    "INGRESO": "Ventas",
    "COSTO_VAR": "Costo de Venta",
    "COSTO_FIJO": "Costo Fijo",
    "OPEX": "OPEX",
    "FINANCIERO": "Gastos Financieros",
    "NO_OPERACIONAL": "Gastos No Operacionales",
}

# Umbral de materialidad por defecto: bajo esto una brecha no se comenta.
UMBRAL_DEFECTO = 1_000_000

# Paleta corporativa (misma de la app)
C_MORADO = "2D0050"
C_MORADO2 = "6B2C91"
C_FUCSIA = "C4007A"
C_VERDE = "0F6E56"
C_ROJO = "CC0000"
C_LILA_BG = "F3E9F7"
C_GRIS = "94A3B8"


# ── CARGA DE DATOS ────────────────────────────────────────────

def cargar_movimientos(ano: int, filtro_soc: str = "", filtro_cc: str = "") -> pd.DataFrame:
    """
    Trae el año completo a nivel periodo × cuenta × centro de costo.
    Se pide el año entero (no solo el YTD) porque el reporte necesita tanto el
    presupuesto comparable de los meses cerrados como el presupuesto anual y el
    de los meses que faltan para la proyección al cierre.
    """
    return query(f"""
        SELECT periodo, sociedad, codigo_cc, nombre_cc, codigo_cuenta, nombre_cuenta,
               clasificacion, categoria_eerr,
               SUM(valor_real) AS real,
               SUM(valor_ppto) AS ppto
        FROM marts.vw_real_vs_ppto
        WHERE periodo BETWEEN :d AND :h {filtro_soc} {filtro_cc}
        GROUP BY periodo, sociedad, codigo_cc, nombre_cc, codigo_cuenta, nombre_cuenta,
                 clasificacion, categoria_eerr
    """, {"d": f"{ano}-01", "h": f"{ano}-12"})


def cargar_por_sociedad(ano: int, filtro_cc: str = "") -> pd.DataFrame:
    """Real y presupuesto por sociedad — sirve para advertir si el presupuesto
    está cargado solo bajo una sociedad (comparación consolidada obligatoria)."""
    return query(f"""
        SELECT sociedad,
               SUM(valor_real) AS real,
               SUM(valor_ppto) AS ppto
        FROM marts.vw_real_vs_ppto
        WHERE periodo BETWEEN :d AND :h {filtro_cc}
        GROUP BY sociedad
        ORDER BY sociedad
    """, {"d": f"{ano}-01", "h": f"{ano}-12"})


def diagnostico_corte(df: pd.DataFrame, ano: int) -> dict:
    """
    Determina hasta qué mes conviene acumular.

    El mes calendario no sirve: puede haber reales cargados a medias (una carga
    en curso). Se compara el ratio real/ppto del último mes con datos contra la
    mediana de los meses anteriores; si cae muy por debajo, el mes se marca como
    parcial y se sugiere el anterior.
    """
    if df.empty:
        return {"ultimo_real": 0, "sugerido": 0, "parcial": False,
                "ratio": 0.0, "ratio_previo": 0.0, "meses_con_real": []}

    m = df.copy()
    m["mes"] = m["periodo"].astype(str).str[5:7].astype(int)
    agg = m.groupby("mes", as_index=False)[["real", "ppto"]].sum()
    con_real = agg[agg["real"] != 0]["mes"].tolist()
    if not con_real:
        return {"ultimo_real": 0, "sugerido": 0, "parcial": False,
                "ratio": 0.0, "ratio_previo": 0.0, "meses_con_real": []}

    ultimo = int(max(con_real))
    ratios = {int(r.mes): (float(r.real) / float(r.ppto)) if r.ppto else None
              for r in agg.itertuples() if int(r.mes) in con_real}
    ratio_ult = ratios.get(ultimo) or 0.0
    previos = [v for k, v in ratios.items() if k < ultimo and v]
    ratio_prev = float(pd.Series(previos).median()) if previos else 0.0

    # Parcial si el último mes ejecutó menos del 60% de lo que ejecutan los meses
    # anteriores en promedio (y hay al menos un mes anterior para comparar).
    parcial = bool(previos) and ratio_ult < 0.60 * ratio_prev
    sugerido = (ultimo - 1) if (parcial and (ultimo - 1) in con_real) else ultimo

    return {"ultimo_real": ultimo, "sugerido": sugerido, "parcial": parcial,
            "ratio": ratio_ult, "ratio_previo": ratio_prev,
            "meses_con_real": sorted(con_real)}


# ── HELPERS DE CÁLCULO ────────────────────────────────────────

def _mes(serie: pd.Series) -> pd.Series:
    return serie.astype(str).str[5:7].astype(int)


def _pct(r: float, p: float) -> float | None:
    """
    % de ejecución (real sobre presupuesto).

    None cuando el presupuesto es cero o negativo: dividir por un presupuesto
    negativo (el plan contempla EBIT negativo en los meses de baja temporada)
    produce porcentajes que no significan nada.
    """
    return (r / p) if p and p > 0 else None


def _clasificar_brecha(r_mes: list, p_mes: list, umbral: float) -> tuple:
    """
    Clasifica el comportamiento de una cuenta a lo largo de los meses cerrados.

    Devuelve (tipo, mes_pico, meses_desviados). El objetivo es distinguir lo que
    un gerente necesita separar: gasto que se repite todos los meses (estructural,
    anualizable), un evento único, o plata que simplemente se movió de mes.
    """
    n = len(r_mes)
    tot_r, tot_p = sum(r_mes), sum(p_mes)
    var = tot_r - tot_p
    difs = [r_mes[i] - p_mes[i] for i in range(n)]
    disp = sum(abs(d) for d in difs)          # varianza bruta, sin compensar
    i_pico = max(range(n), key=lambda i: abs(difs[i])) if n else 0
    mes_pico = i_pico + 1

    if tot_p == 0 and tot_r == 0:
        return "Sin movimiento", mes_pico, 0
    if tot_p == 0:
        return "No presupuestado", mes_pico, sum(1 for d in difs if abs(d) > 0)
    if tot_r == 0:
        return "No ejecutado", mes_pico, n

    if abs(var) < umbral:
        # El acumulado cuadra: si igual hubo vaivén mes a mes, es desfase.
        if disp > max(2 * abs(var), umbral):
            return "Desfase de calendario", mes_pico, sum(
                1 for d in difs if abs(d) > umbral / max(n, 1))
        return "En línea", mes_pico, 0

    conc = (abs(difs[i_pico]) / disp) if disp else 0.0
    signo = 1 if var > 0 else -1
    piso = umbral / max(n, 1)
    meses_desv = sum(1 for d in difs if d * signo > piso)

    if conc >= 0.70:
        return "Puntual", mes_pico, meses_desv
    if meses_desv >= math.ceil(0.60 * n):
        return "Recurrente", mes_pico, meses_desv
    return "Mixto", mes_pico, meses_desv


def _label_sociedad(ac: float, gn: float) -> str:
    """
    De qué sociedad viene el monto: 'ACUÑA 81%', 'Gran Natural 63%' o 'ACUÑA'.

    Siempre lleva el porcentaje de la sociedad que predomina. Nombrarla sola
    sería falso salvo cuando es exclusiva: Administración, por ejemplo, es 81%
    ACUÑA pero tiene $16,5M de Gran Natural repartidos en 21 cuentas. El nombre
    a secas queda reservado para el 100%, que es el único caso en que «solo esa
    sociedad» es cierto.

    Se usa el valor absoluto para el reparto, porque hay cuentas con reversas
    (notas de crédito) que dejarían porcentajes sin sentido con el neto.
    """
    ac, gn = abs(ac), abs(gn)
    total = ac + gn
    if total == 0:
        return "—"
    p_ac = ac / total
    if p_ac >= 0.5:
        dominante, p = ETIQUETA_SOCIEDAD[SOC_ACUNA], p_ac
    else:
        dominante, p = ETIQUETA_SOCIEDAD[SOC_GRAN_NATURAL], 1 - p_ac
    return dominante if p >= 0.995 else f"{dominante} {p*100:.0f}%"


def _texto_meses(pares: list) -> str:
    """
    'Ene +1,8M · Mar +2,1M' — los meses en que la cuenta se salió del
    presupuesto, con el monto de cada uno. Es la respuesta a "¿cuándo pasó?".
    """
    if not pares:
        return "—"
    return " · ".join(
        f"{ABREV_MES[m]} {'+' if v >= 0 else '−'}{abs(v)/1e6:,.1f}M" for m, v in pares)


def _efecto(clasificacion: str, var: float) -> float:
    """
    Impacto de la varianza sobre el resultado.

    En ingresos vender más suma; en cualquier gasto, gastar más resta. Se usa la
    clasificación (y no `signo_presentacion` de dim_cuentas) para que el signo
    sea siempre coherente con la fórmula del EBIT que usa todo el reporte.
    """
    return var if clasificacion == "INGRESO" else -var


# ── CONSTRUCCIÓN DEL REPORTE ──────────────────────────────────

def construir_reporte(df: pd.DataFrame, ano: int, mes_corte: int,
                      sociedad_lbl: str, umbral: float = UMBRAL_DEFECTO,
                      df_soc: pd.DataFrame | None = None,
                      diag: dict | None = None) -> dict:
    """
    Arma todos los bloques del reporte a partir del detalle anual.

    `df` viene de `cargar_movimientos` (año completo). `mes_corte` define el YTD:
    el real se compara contra el presupuesto de esos MISMOS meses, nunca contra
    el presupuesto anual — esa comparación es la que hace ver toda la ejecución
    artificialmente baja.
    """
    d = df.copy()
    d["mes"] = _mes(d["periodo"])
    d["real"] = pd.to_numeric(d["real"], errors="coerce").fillna(0.0)
    d["ppto"] = pd.to_numeric(d["ppto"], errors="coerce").fillna(0.0)

    meses = list(range(1, mes_corte + 1))
    etiquetas_mes = [ABREV_MES[m] for m in meses]
    ytd = d[d["mes"] <= mes_corte]
    resto = d[d["mes"] > mes_corte]

    # Reparto del real entre sociedades: el presupuesto es consolidado, pero el
    # real se factura en una u otra según la cuenta, y saber cuál importa
    # (ACUÑA está en quiebra y se está vaciando hacia Gran Natural).
    def _mix(claves: list, frame: pd.DataFrame) -> pd.DataFrame:
        piv = frame.pivot_table(index=claves, columns="sociedad", values="real",
                                aggfunc="sum", fill_value=0.0)
        for soc in (SOC_ACUNA, SOC_GRAN_NATURAL):
            if soc not in piv.columns:
                piv[soc] = 0.0
        return piv

    mix_cta = _mix(["codigo_cc", "codigo_cuenta"], ytd)
    # El resumen por centro de costo muestra solo gasto, así que su reparto por
    # sociedad tiene que calcularse sobre el mismo universo: incluir las ventas
    # daría un porcentaje que no corresponde al monto de la fila.
    mix_cc = _mix(["codigo_cc"], ytd[ytd["clasificacion"] != "INGRESO"])
    # Para la cabecera se usa la facturación, no el real total: mezclar ventas
    # y gastos en un mismo porcentaje no dice nada. La pregunta de gerencia es
    # cuánto del negocio sigue facturándose en ACUÑA.
    _vta = ytd[ytd["clasificacion"] == "INGRESO"]
    vta_ac = float(_vta.loc[_vta["sociedad"] == SOC_ACUNA, "real"].sum())
    vta_gn = float(_vta.loc[_vta["sociedad"] == SOC_GRAN_NATURAL, "real"].sum())

    # ── 1. P&L YTD por clasificación ──────────────────────────
    def _tot(frame: pd.DataFrame, clasif: str, col: str) -> float:
        return float(frame.loc[frame["clasificacion"] == clasif, col].sum())

    R = {c: _tot(ytd, c, "real") for c in CLASIFS}
    P = {c: _tot(ytd, c, "ppto") for c in CLASIFS}
    PA = {c: _tot(d, c, "ppto") for c in CLASIFS}          # presupuesto anual
    PR = {c: _tot(resto, c, "ppto") for c in CLASIFS}      # presupuesto restante

    def _derivados(x: dict) -> dict:
        ub = x["INGRESO"] - x["COSTO_VAR"]
        ebit = ub - x["COSTO_FIJO"] - x["OPEX"]
        un = ebit - x["FINANCIERO"] - x["NO_OPERACIONAL"]
        return {"UB": ub, "EBIT": ebit, "UN": un}

    DR, DP, DPA, DPR = _derivados(R), _derivados(P), _derivados(PA), _derivados(PR)

    filas_pl = []
    for etiqueta, clasif, sub in LINEAS_PL:
        rv = R[clasif] if clasif else DR[sub]
        pv = P[clasif] if clasif else DP[sub]
        pa = PA[clasif] if clasif else DPA[sub]
        var = rv - pv
        filas_pl.append({
            "Línea": etiqueta,
            "Real YTD": rv,
            "Ppto YTD": pv,
            "Varianza": var,
            "Impacto resultado": _efecto(clasif or "SUBTOTAL", var) if clasif else var,
            "% Ejec.": _pct(rv, pv),
            "% s/Ventas": (rv / R["INGRESO"]) if R["INGRESO"] else None,
            "Ppto Año": pa,
            "% Ppto Año consumido": _pct(rv, pa),
            "_subtotal": sub is not None,
        })
    df_pl = pd.DataFrame(filas_pl)

    margenes = []
    for nombre, clave in [("Margen Bruto", "UB"), ("Margen EBIT", "EBIT"), ("Margen Neto", "UN")]:
        mr = (DR[clave] / R["INGRESO"]) if R["INGRESO"] else 0.0
        mp = (DP[clave] / P["INGRESO"]) if P["INGRESO"] else 0.0
        margenes.append({"Margen": nombre, "Real": mr, "Ppto": mp, "Δ pp": (mr - mp) * 100})
    df_margenes = pd.DataFrame(margenes)

    # ── 2. Puente de EBIT ─────────────────────────────────────
    # Se separan las ventas brutas de otros ingresos y se aísla el efecto del
    # volumen del efecto de la eficiencia del costo variable: si las ventas caen,
    # el costo variable cae con ellas y aparecería como "ahorro" si no se ajusta.
    vb_r = float(ytd.loc[ytd["categoria_eerr"] == "Ventas Brutas", "real"].sum())
    vb_p = float(ytd.loc[ytd["categoria_eerr"] == "Ventas Brutas", "ppto"].sum())
    oi_r = R["INGRESO"] - vb_r
    oi_p = P["INGRESO"] - vb_p
    tasa_cv = (P["COSTO_VAR"] / vb_p) if vb_p else 0.0

    ef_volumen = (vb_r - vb_p) * (1 - tasa_cv)
    ef_cv = -(R["COSTO_VAR"] - vb_r * tasa_cv)
    ef_otros_ing = oi_r - oi_p

    puente = [
        {"Concepto": "EBIT Presupuestado YTD", "Efecto": DP["EBIT"], "Tipo": "inicio"},
        {"Concepto": "Ventas — volumen / precio", "Efecto": ef_volumen, "Tipo": "efecto"},
        {"Concepto": "Costo variable — eficiencia y mix", "Efecto": ef_cv, "Tipo": "efecto"},
    ]
    if abs(ef_otros_ing) > 0:
        puente.append({"Concepto": "Otros ingresos", "Efecto": ef_otros_ing, "Tipo": "efecto"})

    ccs = sorted(set(d["codigo_cc"].dropna()) - {"CC-00"})
    for clasif, etiqueta in [("COSTO_FIJO", "Costo fijo"), ("OPEX", "OPEX")]:
        for cc in ccs:
            sub = ytd[(ytd["clasificacion"] == clasif) & (ytd["codigo_cc"] == cc)]
            v = float(sub["real"].sum() - sub["ppto"].sum())
            if abs(v) < 1:
                continue
            puente.append({"Concepto": f"{etiqueta} — {NOMBRES_CC.get(cc, cc)}",
                           "Efecto": -v, "Tipo": "efecto"})
        # Gasto de la clasificación que no cuelga de un CC operativo
        sub0 = ytd[(ytd["clasificacion"] == clasif) & (~ytd["codigo_cc"].isin(ccs))]
        v0 = float(sub0["real"].sum() - sub0["ppto"].sum())
        if abs(v0) >= 1:
            puente.append({"Concepto": f"{etiqueta} — sin centro de costo",
                           "Efecto": -v0, "Tipo": "efecto"})

    puente.append({"Concepto": "EBIT Real YTD", "Efecto": DR["EBIT"], "Tipo": "fin"})
    df_puente = pd.DataFrame(puente)

    # Control de cuadratura: los efectos deben llevar del EBIT ppto al EBIT real.
    suma_efectos = float(df_puente.loc[df_puente["Tipo"] == "efecto", "Efecto"].sum())
    descuadre = DP["EBIT"] + suma_efectos - DR["EBIT"]

    # ── 3. Centros de costo ───────────────────────────────────
    gasto_ytd = ytd[ytd["clasificacion"] != "INGRESO"]
    gasto_anual = d[d["clasificacion"] != "INGRESO"]
    filas_cc = []
    for cc in sorted(set(gasto_ytd["codigo_cc"].dropna())):
        s = gasto_ytd[gasto_ytd["codigo_cc"] == cc]
        rv, pv = float(s["real"].sum()), float(s["ppto"].sum())
        pa = float(gasto_anual.loc[gasto_anual["codigo_cc"] == cc, "ppto"].sum())
        pr = float(gasto_anual.loc[(gasto_anual["codigo_cc"] == cc) &
                                   (gasto_anual["mes"] > mes_corte), "ppto"].sum())
        run_rate = (rv / mes_corte * 12) if mes_corte else 0.0
        filas_cc.append({
            "Centro de costo": _label_cc(cc, s["nombre_cc"].iloc[0] if len(s) else cc),
            "Código": cc,
            "Sociedad": (_label_sociedad(mix_cc.loc[cc, SOC_ACUNA],
                                         mix_cc.loc[cc, SOC_GRAN_NATURAL])
                         if cc in mix_cc.index else "—"),
            "Real YTD": rv,
            "Ppto YTD": pv,
            "Varianza": rv - pv,
            "Impacto resultado": -(rv - pv),
            "% Ejec.": _pct(rv, pv),
            "Ppto Año": pa,
            "% Ppto Año consumido": _pct(rv, pa),
            "Proyección cierre": rv + pr,
            "Run-rate x12": run_rate,
            "Desv. proyectada": run_rate - pa,
        })
    df_cc = pd.DataFrame(filas_cc).sort_values("Código").reset_index(drop=True)

    # Apertura CC × línea del P&L
    filas_ccl = []
    for cc in sorted(set(gasto_ytd["codigo_cc"].dropna())):
        for clasif in ["COSTO_FIJO", "OPEX", "FINANCIERO", "NO_OPERACIONAL"]:
            s = gasto_ytd[(gasto_ytd["codigo_cc"] == cc) & (gasto_ytd["clasificacion"] == clasif)]
            if s.empty or (s["real"].sum() == 0 and s["ppto"].sum() == 0):
                continue
            rv, pv = float(s["real"].sum()), float(s["ppto"].sum())
            filas_ccl.append({
                "Centro de costo": _label_cc(cc),
                "Línea": {"COSTO_FIJO": "Costo Fijo", "OPEX": "OPEX",
                          "FINANCIERO": "Gastos Financieros",
                          "NO_OPERACIONAL": "Gastos No Operacionales"}[clasif],
                "Real YTD": rv, "Ppto YTD": pv, "Varianza": rv - pv,
                "Impacto resultado": -(rv - pv), "% Ejec.": _pct(rv, pv),
            })
    df_cc_linea = pd.DataFrame(filas_ccl)

    # ── 4. Detalle por cuenta × centro de costo ───────────────
    piv_r = ytd.pivot_table(index=["codigo_cc", "codigo_cuenta"], columns="mes",
                            values="real", aggfunc="sum", fill_value=0.0)
    piv_p = ytd.pivot_table(index=["codigo_cc", "codigo_cuenta"], columns="mes",
                            values="ppto", aggfunc="sum", fill_value=0.0)
    meta = (d.groupby(["codigo_cc", "codigo_cuenta"])
              .agg(nombre_cuenta=("nombre_cuenta", "first"),
                   nombre_cc=("nombre_cc", "first"),
                   clasificacion=("clasificacion", "first"),
                   categoria_eerr=("categoria_eerr", "first"),
                   ppto_anual=("ppto", "sum"))
              .reset_index())

    filas_det = []
    for idx in piv_r.index.union(piv_p.index):
        r_mes = [float(piv_r.loc[idx, m]) if (idx in piv_r.index and m in piv_r.columns) else 0.0
                 for m in meses]
        p_mes = [float(piv_p.loc[idx, m]) if (idx in piv_p.index and m in piv_p.columns) else 0.0
                 for m in meses]
        rv, pv = sum(r_mes), sum(p_mes)
        if rv == 0 and pv == 0:
            continue
        cc, cuenta = idx
        info = meta[(meta["codigo_cc"] == cc) & (meta["codigo_cuenta"] == cuenta)]
        if info.empty:
            continue
        info = info.iloc[0]
        var = rv - pv
        clasif = info["clasificacion"]
        tipo, mes_pico, n_desv = _clasificar_brecha(r_mes, p_mes, umbral)
        impacto = _efecto(clasif, var)
        # Solo se anualizan los gastos: extrapolar ventas x12 en un negocio
        # estacional (verano) daría una cifra sin sentido.
        anualizado = (var / mes_corte * 12) if (
            tipo == "Recurrente" and mes_corte and clasif != "INGRESO") else None
        pr_cta = float(d.loc[(d["codigo_cc"] == cc) & (d["codigo_cuenta"] == cuenta) &
                             (d["mes"] > mes_corte), "ppto"].sum())

        # Detalle mes a mes de la brecha: en qué meses se pasó del presupuesto
        # y por cuánto. Sin esto la brecha anual no se puede accionar.
        difs = [r_mes[i] - p_mes[i] for i in range(len(meses))]
        piso = umbral / max(mes_corte, 1)
        exceso = [(meses[i], difs[i]) for i in range(len(meses))
                  if _efecto(clasif, difs[i]) < -piso]

        fila = {
            "Centro de costo": _label_cc(cc, info["nombre_cc"]),
            "Código CC": cc,
            "Cuenta": f"{cuenta} {info['nombre_cuenta']}",
            "Sociedad": (_label_sociedad(mix_cta.loc[idx, SOC_ACUNA],
                                         mix_cta.loc[idx, SOC_GRAN_NATURAL])
                         if idx in mix_cta.index else "—"),
            "Línea P&L": ETIQUETA_LINEA.get(clasif, clasif),
            "_clasif": clasif,
            "Categoría EERR": info["categoria_eerr"],
            "Real YTD": rv,
            "Ppto YTD": pv,
            "Varianza": var,
            "Impacto resultado": impacto,
            "% Ejec.": _pct(rv, pv),
            "Tipo de brecha": tipo,
            "Meses con desvío desfavorable": _texto_meses(exceso),
            "Meses desviados": n_desv,
            "Mes pico": ABREV_MES.get(mes_pico, ""),
            "Impacto anualizado": anualizado,
            "Ppto Año": float(info["ppto_anual"]),
            "Proyección cierre": rv + pr_cta,
            "_meses_exceso": exceso,
            "_real_mes": r_mes,
            "_ppto_mes": p_mes,
        }
        # Varianza de cada mes en columnas propias (real − presupuesto)
        for i, mm in enumerate(meses):
            fila[f"Var {ABREV_MES[mm]}"] = difs[i]
        filas_det.append(fila)

    df_det = pd.DataFrame(filas_det)
    if not df_det.empty:
        df_det = (df_det.reindex(df_det["Impacto resultado"].abs()
                                 .sort_values(ascending=False).index)
                  .reset_index(drop=True))

    # ── 5. Plan de acción: Pareto de lo desfavorable ──────────
    if not df_det.empty:
        desfav = df_det[(df_det["Impacto resultado"] < 0) &
                        (df_det["Impacto resultado"].abs() >= umbral)].copy()
    else:
        desfav = pd.DataFrame()

    if not desfav.empty:
        desfav = desfav.sort_values("Impacto resultado")
        total_desfav = float(-desfav["Impacto resultado"].sum())
        desfav["_acum"] = (-desfav["Impacto resultado"]).cumsum()
        desfav["% acum. brecha"] = desfav["_acum"] / total_desfav if total_desfav else 0.0
        # Corte Pareto: hasta explicar el 80% de la brecha desfavorable
        corte = desfav[desfav["_acum"] <= 0.80 * total_desfav]
        n_pareto = max(len(corte) + 1, min(5, len(desfav)))
        df_accion = desfav.head(n_pareto).copy()
    else:
        total_desfav = 0.0
        df_accion = pd.DataFrame()

    cols_var_mes = [f"Var {ABREV_MES[m]}" for m in meses]
    if not df_accion.empty:
        df_accion = df_accion[[
            "Centro de costo", "Cuenta", "Sociedad", "Línea P&L", "Real YTD", "Ppto YTD",
            "Varianza", "Impacto resultado", "% acum. brecha", "Tipo de brecha",
            "Meses con desvío desfavorable", "Meses desviados", "Mes pico", "Impacto anualizado",
            *cols_var_mes,
        ] + ["_meses_exceso", "_real_mes", "_ppto_mes"]].reset_index(drop=True)
        for col in ["Explicación (qué pasó)", "Acción comprometida", "Responsable", "Fecha"]:
            df_accion[col] = ""

    # ── 6. Mes a mes ──────────────────────────────────────────
    filas_mm = []
    for cc in sorted(set(gasto_ytd["codigo_cc"].dropna())):
        for concepto, col in [("Real", "real"), ("Ppto", "ppto")]:
            fila = {"Centro de costo": _label_cc(cc), "Concepto": concepto}
            s = gasto_ytd[gasto_ytd["codigo_cc"] == cc]
            for m in meses:
                fila[ABREV_MES[m]] = float(s.loc[s["mes"] == m, col].sum())
            fila["Total YTD"] = sum(fila[ABREV_MES[m]] for m in meses)
            filas_mm.append(fila)
        fila_v = {"Centro de costo": _label_cc(cc), "Concepto": "Varianza"}
        s = gasto_ytd[gasto_ytd["codigo_cc"] == cc]
        for m in meses:
            fila_v[ABREV_MES[m]] = float(s.loc[s["mes"] == m, "real"].sum() -
                                         s.loc[s["mes"] == m, "ppto"].sum())
        fila_v["Total YTD"] = sum(fila_v[ABREV_MES[m]] for m in meses)
        filas_mm.append(fila_v)
    df_mes_cc = pd.DataFrame(filas_mm)

    filas_plm = []
    for etiqueta, clasif, sub in LINEAS_PL:
        for concepto, col in [("Real", "real"), ("Ppto", "ppto")]:
            fila = {"Línea": etiqueta, "Concepto": concepto}
            for m in meses:
                mm = ytd[ytd["mes"] == m]
                if clasif:
                    v = float(mm.loc[mm["clasificacion"] == clasif, col].sum())
                else:
                    base = {c: float(mm.loc[mm["clasificacion"] == c, col].sum()) for c in CLASIFS}
                    v = _derivados(base)[sub]
                fila[ABREV_MES[m]] = v
            fila["Total YTD"] = sum(fila[ABREV_MES[m]] for m in meses)
            filas_plm.append(fila)
    df_mes_pl = pd.DataFrame(filas_plm)

    # Mes a mes de las cuentas con brecha material: Real / Ppto / Varianza por
    # mes. Es el respaldo de "en qué mes se pasó" para cada cuenta comentada.
    filas_ctam = []
    if not df_det.empty:
        materiales = df_det[df_det["Impacto resultado"].abs() >= umbral]
        for _, r in materiales.iterrows():
            for concepto, serie in [("Real", r["_real_mes"]), ("Ppto", r["_ppto_mes"]),
                                    ("Varianza", [r["_real_mes"][i] - r["_ppto_mes"][i]
                                                  for i in range(len(meses))])]:
                fila = {"Centro de costo": r["Centro de costo"],
                        "Cuenta": r["Cuenta"], "Concepto": concepto}
                for i, mm in enumerate(meses):
                    fila[ABREV_MES[mm]] = float(serie[i])
                fila["Total YTD"] = float(sum(serie))
                filas_ctam.append(fila)
    df_mes_cuenta = pd.DataFrame(filas_ctam)

    # ── 7. Proyección al cierre ───────────────────────────────
    filas_proy = []
    for etiqueta, clasif, sub in LINEAS_PL:
        rv = R[clasif] if clasif else DR[sub]
        pa = PA[clasif] if clasif else DPA[sub]
        pr = PR[clasif] if clasif else DPR[sub]
        run = (rv / mes_corte * 12) if mes_corte else 0.0
        proy = rv + pr
        filas_proy.append({
            "Línea": etiqueta,
            "Real YTD": rv,
            "Ppto restante": pr,
            "Proyección (real + ppto restante)": proy,
            "Run-rate x12": run,
            "Ppto Año": pa,
            "Desv. vs Ppto Año": proy - pa,
            "_subtotal": sub is not None,
        })
    df_proy = pd.DataFrame(filas_proy)

    # ── 8. Bases y alertas ────────────────────────────────────
    alertas = []
    if df_soc is not None and not df_soc.empty:
        s = df_soc.copy()
        s["real"] = pd.to_numeric(s["real"], errors="coerce").fillna(0.0)
        s["ppto"] = pd.to_numeric(s["ppto"], errors="coerce").fillna(0.0)
        sin_ppto = s[(s["ppto"] == 0) & (s["real"] > 0)]["sociedad"].tolist()
        if sin_ppto:
            con_ppto = s[s["ppto"] > 0]["sociedad"].tolist()
            alertas.append(
                f"El presupuesto {ano} está cargado íntegramente bajo "
                f"{', '.join(con_ppto) or '—'}. {', '.join(sin_ppto)} registra real "
                f"sin presupuesto propio, por lo que la comparación válida es la "
                f"CONSOLIDADA (ambas sociedades): la operación se factura en una u "
                f"otra según el mes. Filtrar por una sola sociedad rompe la comparación."
            )
    corte_parcial = bool(diag and diag.get("parcial")
                         and mes_corte >= diag.get("ultimo_real", 0))
    if diag and diag.get("parcial"):
        estado = (
            f"{ABREV_MES.get(diag['ultimo_real'], '')} tiene reales cargados parcialmente "
            f"({diag['ratio']*100:.0f}% del presupuesto del mes, contra "
            f"{diag['ratio_previo']*100:.0f}% típico de los meses cerrados). "
        )
        if corte_parcial:
            alertas.append(
                estado +
                f"Este reporte INCLUYE ese mes: los ahorros y la subejecución de gastos "
                f"están sobrestimados, porque hay facturas del mes todavía sin cargar. "
                f"Para una lectura firme, cortar en "
                f"{ABREV_MES.get(diag['sugerido'], '')}; lo que ya aparece SOBRE "
                f"presupuesto en {ABREV_MES.get(diag['ultimo_real'], '')} sí es real "
                f"(solo puede aumentar al completarse la carga)."
            )
        else:
            alertas.append(
                estado +
                f"El corte se fijó en {ABREV_MES.get(mes_corte, '')} para no mostrar "
                f"ahorros que en realidad son facturas por cargar."
            )
    if abs(descuadre) > 1:
        alertas.append(f"Descuadre en el puente de EBIT: ${descuadre:,.0f}. Revisar clasificaciones.")

    if not df_det.empty:
        # Solo gastos: un ingreso sin presupuesto no es un problema de control
        # presupuestario, es una venta que no estaba en el plan.
        sin_ppto = df_det[(df_det["Tipo de brecha"] == "No presupuestado") &
                          (df_det["_clasif"] != "INGRESO")]
        if not sin_ppto.empty:
            monto = float(sin_ppto["Real YTD"].sum())
            alertas.append(
                f"{_plural(len(sin_ppto), 'cuenta de gasto', 'cuentas de gasto')} con real "
                f"y sin presupuesto asignado (${monto/1e6:,.1f}M YTD). No tienen contra qué "
                f"medirse: requieren presupuesto o reclasificación."
            )

    # El reporte mezcla tres escalas a propósito (acumulada, mensual y anual):
    # se explicitan para que nadie compare columnas de escalas distintas.
    escalas = [
        f"ACUMULADO Enero–{ABREV_MES.get(mes_corte, '')} ({mes_corte} de 12 meses) — columnas "
        f"«Real YTD», «Ppto YTD», «Varianza», «Impacto resultado» y «% Ejec.», más las hojas "
        f"1 Resumen Ejecutivo y 2 Puente EBIT. El real se compara siempre contra el "
        f"presupuesto de esos mismos meses.",
        "VALOR DE CADA MES (no acumulado) — toda la hoja «6 Mes a Mes» y las columnas "
        "«Var <mes>» de las hojas 4 y 5. En la hoja 6, la columna «Total YTD» es la suma "
        "de esos meses y cuadra con el «Real YTD» de las demás hojas.",
        "AÑO COMPLETO (12 meses) — «Ppto Año», «% Ppto Año consumido», «Proyección cierre», "
        "«Run-rate x12» y «Desv. proyectada», más la hoja 7 Proyección Cierre.",
        "Las tres escalas conviven a propósito: un centro de costo puede ir sobre-ejecutado "
        f"contra el presupuesto de {mes_corte} meses y, a la vez, llevar consumido menos de "
        f"la mitad de su presupuesto anual. Ambas cifras son correctas y miden cosas distintas.",
    ]

    bases = [
        f"Fuente: marts.vw_real_vs_ppto (real desde ERP Obuma; presupuesto editable en la app).",
        f"El real acumulado se compara contra el presupuesto de los MISMOS meses "
        f"(Ene–{ABREV_MES.get(mes_corte, '')}), no contra el presupuesto anual.",
        f"Umbral de materialidad aplicado: ${umbral/1e6:,.1f}M por cuenta y centro de costo.",
        "Signo: en gastos, varianza positiva = se gastó más que el presupuesto. "
        "La columna «Impacto resultado» ya viene con el signo del efecto sobre el EBIT.",
        "Parte del costo de fábrica está en el costo fijo de Producción, no en el costo "
        "variable: el margen bruto de este reporte es un margen de contribución.",
        "El costo variable proviene de carga manual (staging.cv_real_manual).",
        "Cifras en pesos chilenos. El formato «M» muestra millones; el valor de la celda "
        "está en pesos y es apto para pivotear.",
    ]

    # ── 9. Conclusiones calculadas ────────────────────────────
    conclusiones = _conclusiones(
        R, P, DR, DP, PA, DPA, DPR, mes_corte, ef_volumen, ef_cv, ef_otros_ing,
        df_cc, df_det, umbral, ano, corte_parcial, diag)

    _tot_soc = abs(vta_ac) + abs(vta_gn)
    mix_lbl = (f"{ETIQUETA_SOCIEDAD[SOC_ACUNA]} {vta_ac/1e6:,.0f}M "
               f"({abs(vta_ac)/_tot_soc*100:.0f}%) · "
               f"{ETIQUETA_SOCIEDAD[SOC_GRAN_NATURAL]} {vta_gn/1e6:,.0f}M "
               f"({abs(vta_gn)/_tot_soc*100:.0f}%)") if _tot_soc else "—"

    return {
        "meta": {
            "ano": ano, "mes_corte": mes_corte, "n_meses": mes_corte,
            "mix_sociedad": mix_lbl,
            "mes_corte_nombre": ABREV_MES.get(mes_corte, ""),
            "periodo_lbl": f"Enero–{ABREV_MES.get(mes_corte, '')} {ano}",
            "sociedad": sociedad_lbl,
            "umbral": umbral,
            "generado": datetime.now().strftime("%d-%m-%Y %H:%M"),
            "meses": etiquetas_mes,
        },
        "kpi": {
            "ventas_r": R["INGRESO"], "ventas_p": P["INGRESO"],
            "ub_r": DR["UB"], "ub_p": DP["UB"],
            "ebit_r": DR["EBIT"], "ebit_p": DP["EBIT"],
            "un_r": DR["UN"], "un_p": DP["UN"],
            "ebit_proy": DR["EBIT"] + DPR["EBIT"], "ebit_ppto_ano": DPA["EBIT"],
            "gasto_r": sum(R[c] for c in CLASIFS if c != "INGRESO"),
            "gasto_p": sum(P[c] for c in CLASIFS if c != "INGRESO"),
        },
        "pl": df_pl,
        "margenes": df_margenes,
        "puente": df_puente,
        "descuadre": descuadre,
        "cc": df_cc,
        "cc_linea": df_cc_linea,
        "detalle": df_det,
        "accion": df_accion,
        "total_desfavorable": total_desfav,
        "mes_cc": df_mes_cc,
        "mes_pl": df_mes_pl,
        "mes_cuenta": df_mes_cuenta,
        "cols_var_mes": cols_var_mes,
        "proyeccion": df_proy,
        "alertas": alertas,
        "escalas": escalas,
        "bases": bases,
        "conclusiones": conclusiones,
    }


def _conclusiones(R, P, DR, DP, PA, DPA, DPR, mes_corte, ef_vol, ef_cv, ef_oi,
                  df_cc, df_det, umbral, ano, corte_parcial=False, diag=None) -> list:
    """Redacta los hallazgos a partir de las cifras. Reglas deterministas: el
    mismo dato produce siempre el mismo texto (no interviene la IA)."""
    def m(v):
        return f"${v/1e6:,.1f}M"

    out = []
    if not any(P.values()):
        # Selección sin presupuesto cargado (p. ej. una sociedad sola): no hay
        # comparación posible, decirlo antes que cualquier otra cifra.
        out.append(
            f"Esta selección no tiene presupuesto cargado para el periodo: el EBIT real "
            f"acumulado es {m(DR['EBIT'])} y no hay contra qué medirlo. Para el control "
            f"presupuestario usa la vista consolidada."
        )
        return out

    if corte_parcial and diag:
        # El corte incluye un mes a medio cargar: la advertencia va primero,
        # porque teñe todas las cifras de gasto que vienen después.
        out.append(
            f"LECTURA CONDICIONADA: el acumulado incluye {ABREV_MES.get(diag['ultimo_real'], '')}, "
            f"que está cargado parcialmente ({diag['ratio']*100:.0f}% de su presupuesto contra "
            f"{diag['ratio_previo']*100:.0f}% típico). Los ahorros de gasto que se leen abajo "
            f"están sobrestimados y bajarán al completarse la carga. Lo que ya aparece SOBRE "
            f"presupuesto en ese mes sí es firme: solo puede subir."
        )

    gap_ebit = DR["EBIT"] - DP["EBIT"]
    signo = "sobre" if gap_ebit >= 0 else "bajo"
    if DP["EBIT"] > 0:
        cierre = f"({DR['EBIT'] / DP['EBIT'] * 100:.0f}% de cumplimiento)"
    else:
        # El presupuesto de estos meses ya contemplaba pérdida operacional: el
        # porcentaje de cumplimiento no aporta, la lectura es la brecha en pesos.
        cierre = ("— el presupuesto de estos meses ya contemplaba resultado operacional "
                  "negativo, propio de la estacionalidad del negocio")
    out.append(
        f"EBIT acumulado de {m(DR['EBIT'])} contra {m(DP['EBIT'])} presupuestados "
        f"para los mismos {mes_corte} meses: {m(abs(gap_ebit))} {signo} el objetivo "
        f"{cierre}."
    )

    # Qué explica la brecha
    gastos = [(r["Centro de costo"], float(r["Impacto resultado"])) for _, r in df_cc.iterrows()]
    ef_gastos = sum(v for _, v in gastos)
    piezas = [("mayores/menores ventas", ef_vol), ("costo variable", ef_cv),
              ("otros ingresos", ef_oi), ("gastos por centro de costo", ef_gastos)]
    piezas = sorted([p for p in piezas if abs(p[1]) >= umbral],
                    key=lambda x: abs(x[1]), reverse=True)
    if piezas:
        detalle = "; ".join(
            f"{nombre} {'+' if v >= 0 else '−'}{m(abs(v))}" for nombre, v in piezas)
        out.append(f"Descomposición de la brecha de EBIT: {detalle}.")

    # Centro de costo que más pesa (CC-00 no es un centro de costo: son las
    # ventas y el costo variable, ya explicados en la descomposición anterior)
    cc_op = df_cc[df_cc["Código"] != CC_SIN_ASIGNAR] if not df_cc.empty else df_cc
    if not cc_op.empty:
        cc_ord = cc_op.reindex(cc_op["Impacto resultado"].abs().sort_values(ascending=False).index)
        top_cc = cc_ord.iloc[0]
        imp = float(top_cc["Impacto resultado"])
        sentido = "sobregasto" if imp < 0 else "ahorro"
        pct_ej = top_cc["% Ejec."]
        pct_txt = f"{pct_ej*100:.0f}% de ejecución" if pd.notna(pct_ej) and pct_ej is not None else "sin presupuesto"
        out.append(
            f"{top_cc['Centro de costo']} es el centro de costo con mayor desvío: "
            f"{m(abs(imp))} de {sentido} ({pct_txt}); consumió el "
            f"{(top_cc['% Ppto Año consumido'] or 0)*100:.0f}% de su presupuesto anual "
            f"con {mes_corte} de 12 meses."
        )

    # Estructural vs desfase
    if not df_det.empty:
        desf = df_det[df_det["Impacto resultado"] < -umbral]
        rec = desf[desf["Tipo de brecha"] == "Recurrente"]
        pun = desf[desf["Tipo de brecha"] == "Puntual"]
        nop = desf[desf["Tipo de brecha"] == "No presupuestado"]
        if not desf.empty:
            partes = []
            if not rec.empty:
                partes.append(
                    f"{m(abs(rec['Impacto resultado'].sum()))} es sobregasto recurrente "
                    f"({_plural(len(rec), 'cuenta', 'cuentas')}, "
                    f"{m(abs(rec['Impacto anualizado'].sum()))} de impacto si se mantiene "
                    f"los 12 meses)")
            if not pun.empty:
                partes.append(f"{m(abs(pun['Impacto resultado'].sum()))} son eventos puntuales "
                              f"({_plural(len(pun), 'cuenta', 'cuentas')})")
            if not nop.empty:
                partes.append(f"{m(abs(nop['Impacto resultado'].sum()))} corresponde a gasto "
                              f"no presupuestado ({_plural(len(nop), 'cuenta', 'cuentas')})")
            if partes:
                out.append("Del desvío desfavorable acumulado: " + "; ".join(partes) + ".")

        # Top 3 cuentas
        top3 = df_det[df_det["Impacto resultado"] < 0].head(3)
        if not top3.empty:
            items = "; ".join(
                f"{r['Cuenta']} ({r['Centro de costo']}) {m(abs(r['Impacto resultado']))} — "
                f"{r['Tipo de brecha'].lower()}"
                + (f", con desvío en contra en {r['Meses con desvío desfavorable']}"
                   if r["Meses con desvío desfavorable"] != "—" else "")
                for _, r in top3.iterrows())
            out.append(f"Cuentas que más restan al resultado: {items}.")

        desfase = df_det[df_det["Tipo de brecha"] == "Desfase de calendario"]
        if not desfase.empty:
            out.append(
                f"{_plural(len(desfase), 'cuenta muestra', 'cuentas muestran')} desfase de "
                f"calendario: el acumulado cuadra con el presupuesto pero el gasto cayó en "
                f"meses distintos a los planificados. No son ahorro ni sobregasto — es "
                f"calendario, no monto — y conviene corregir la mensualización del presupuesto."
            )

    # Ventas y ritmo del año
    pct_v = (R["INGRESO"] / P["INGRESO"] * 100) if P["INGRESO"] else 0
    pct_anual = (R["INGRESO"] / PA["INGRESO"] * 100) if PA["INGRESO"] else 0
    out.append(
        f"Ventas YTD {m(R['INGRESO'])}: {pct_v:.0f}% del presupuesto de los meses "
        f"transcurridos y {pct_anual:.0f}% del presupuesto anual, con {mes_corte}/12 "
        f"meses ({mes_corte/12*100:.0f}% del año). El negocio es estacional: la "
        f"comparación válida es la de los meses cerrados."
    )

    # Proyección
    ebit_proy = DR["EBIT"] + DPR["EBIT"]
    if DPA["EBIT"]:
        out.append(
            f"Si los meses restantes se ejecutan según presupuesto, el EBIT cierra en "
            f"{m(ebit_proy)} contra {m(DPA['EBIT'])} presupuestados "
            f"({m(ebit_proy - DPA['EBIT'])} de desviación al cierre)."
        )
    return out


# ── EXPORT EXCEL ──────────────────────────────────────────────

_FMT_M = '#,##0.0,, "M"'
_FMT_PCT = "0.0%"
_FMT_PP = '+0.0" pp";-0.0" pp"'
_FMT_INT = "#,##0"


class _Hoja:
    """Escritor de hojas con el estilo corporativo (encabezados morados,
    subtotales resaltados, cifras en millones)."""

    def __init__(self, wb, nombre: str, ancho_col_a: int = 34):
        from openpyxl.utils import get_column_letter
        self.ws = wb.create_sheet(title=nombre[:31])
        self.r = 1
        self.ancho_col_a = ancho_col_a
        self._gcl = get_column_letter
        self.max_col = 1
        self.fila_cabecera = None   # fila de encabezados de la primera tabla

    def titulo(self, texto: str, sub: str = ""):
        from openpyxl.styles import Font
        c = self.ws.cell(row=self.r, column=1, value=texto)
        c.font = Font(bold=True, size=14, color=C_MORADO)
        self.r += 1
        if sub:
            c = self.ws.cell(row=self.r, column=1, value=sub)
            c.font = Font(size=10, italic=True, color=C_MORADO2)
            self.r += 1
        self.r += 1

    def seccion(self, texto: str):
        from openpyxl.styles import Font
        c = self.ws.cell(row=self.r, column=1, value=texto)
        c.font = Font(bold=True, size=12, color=C_MORADO)
        self.r += 1

    def texto(self, lineas: list, vinetas: bool = True, color: str = "333333"):
        from openpyxl.styles import Font, Alignment
        for t in lineas:
            c = self.ws.cell(row=self.r, column=1, value=(f"•  {t}" if vinetas else t))
            c.font = Font(size=10, color=color)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            self.ws.row_dimensions[self.r].height = max(15, 13 * (1 + len(str(t)) // 110))
            self.r += 1
        self.r += 1

    def tabla(self, df: pd.DataFrame, formatos: dict | None = None,
              subtotales: list | None = None, resaltar_signo: list | None = None,
              anchos: dict | None = None):
        """Escribe un DataFrame con encabezado morado. `formatos` mapea columna →
        número de formato; `resaltar_signo` colorea verde/rojo según el signo."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        if df is None or df.empty:
            self.texto(["Sin datos para este bloque con los filtros aplicados."])
            return

        cols = [c for c in df.columns if not str(c).startswith("_")]
        formatos = formatos or {}
        resaltar_signo = resaltar_signo or []
        borde = Border(bottom=Side(style="thin", color="E6DCEF"))
        fill_h = PatternFill("solid", fgColor=C_MORADO2)
        fill_sub = PatternFill("solid", fgColor=C_LILA_BG)

        if self.fila_cabecera is None:
            self.fila_cabecera = self.r
        for j, col in enumerate(cols, start=1):
            c = self.ws.cell(row=self.r, column=j, value=str(col))
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = fill_h
            c.alignment = Alignment(horizontal="left" if j == 1 else "right",
                                    wrap_text=True, vertical="center")
        self.ws.row_dimensions[self.r].height = 28
        self.r += 1

        for i, (_, fila) in enumerate(df.iterrows()):
            es_sub = bool(subtotales and i in subtotales)
            for j, col in enumerate(cols, start=1):
                v = fila[col]
                if pd.isna(v):
                    v = None
                elif hasattr(v, "item"):
                    v = v.item()
                c = self.ws.cell(row=self.r, column=j, value=v)
                c.border = borde
                c.font = Font(size=10, bold=es_sub,
                              color=C_MORADO if es_sub else "222222")
                if es_sub:
                    c.fill = fill_sub
                if col in formatos:
                    c.number_format = formatos[col]
                if col in resaltar_signo and isinstance(v, (int, float)):
                    c.font = Font(size=10, bold=es_sub,
                                  color=(C_VERDE if v >= 0 else C_ROJO))
            self.r += 1
        self.r += 1
        self.max_col = max(self.max_col, len(cols))

        self.ws.column_dimensions["A"].width = self.ancho_col_a
        for j in range(2, len(cols) + 1):
            letra = self._gcl(j)
            ancho = (anchos or {}).get(cols[j - 1], 15)
            actual = self.ws.column_dimensions[letra].width or 0
            self.ws.column_dimensions[letra].width = max(actual, ancho)

    def congelar(self, columna: str = "A"):
        """Deja fijos los encabezados de la primera tabla y las columnas a su
        izquierda, para que al desplazarse no se pierda de vista qué se está mirando."""
        fila = (self.fila_cabecera + 1) if self.fila_cabecera else 1
        self.ws.freeze_panes = f"{columna}{fila}"


def to_excel(rep: dict) -> bytes:
    """Genera el workbook completo del reporte de gerencia."""
    from openpyxl import Workbook

    meta = rep["meta"]
    kpi = rep["kpi"]
    wb = Workbook()
    wb.remove(wb.active)

    cab = f"Kreems · Control Presupuestario {meta['ano']}"
    sub = f"{meta['sociedad']} · {meta['periodo_lbl']} · generado {meta['generado']}"

    # ── Portada ───────────────────────────────────────────────
    h = _Hoja(wb, "Portada", ancho_col_a=118)
    h.titulo("Reporte de Gerencia — Real vs Presupuesto", cab)
    h.seccion("Alcance")
    h.texto([f"Sociedad: {meta['sociedad']}",
             f"Periodo acumulado: {meta['periodo_lbl']} ({meta['n_meses']} de 12 meses)",
             f"Facturación del periodo: {meta['mix_sociedad']}",
             "El presupuesto es uno solo para el negocio. La columna «Sociedad» de las hojas "
             "3, 4 y 5 indica dónde se registra cada línea: nombra la sociedad que concentra "
             "el mayor monto y con qué porcentaje (ej. «ACUÑA 81%» = el 19% restante está en "
             "Gran Natural). Sin porcentaje = el 100% está en esa sociedad.",
             f"Generado: {meta['generado']}"])
    if rep["alertas"]:
        h.seccion("Advertencias sobre la base de comparación")
        h.texto(rep["alertas"])
    h.seccion("Conclusiones")
    h.texto(rep["conclusiones"])
    h.seccion("Cómo leer las columnas")
    h.texto(rep["escalas"])
    h.seccion("Bases y criterios")
    h.texto(rep["bases"])
    h.seccion("Contenido")
    h.texto([
        "1 Resumen Ejecutivo — P&L acumulado contra el presupuesto de los mismos meses.",
        "2 Puente EBIT — de dónde sale la diferencia entre el EBIT presupuestado y el real.",
        "3 Centros de Costo — ejecución, consumo del presupuesto anual y proyección por CC.",
        "4 Detalle Cuentas — cada cuenta con su tipo de brecha (recurrente / puntual / desfase).",
        "5 Plan de Acción — las desviaciones que explican el 80% de la brecha, para completar en la reunión.",
        "6 Mes a Mes — perfil mensual por centro de costo, por línea del P&L y por cuenta "
        "con brecha material (en qué mes exactamente se produjo cada desviación).",
        "7 Proyección Cierre — cómo termina el año si los meses restantes se ejecutan según presupuesto.",
    ])

    # ── 1. Resumen ejecutivo ──────────────────────────────────
    h = _Hoja(wb, "1 Resumen Ejecutivo", ancho_col_a=26)
    h.titulo("Resumen Ejecutivo", sub)
    df_pl = rep["pl"]
    subs = [i for i, v in enumerate(df_pl["_subtotal"]) if v]
    h.seccion(f"Estado de Resultados acumulado — {meta['periodo_lbl']}")
    h.tabla(df_pl,
            formatos={"Real YTD": _FMT_M, "Ppto YTD": _FMT_M, "Varianza": _FMT_M,
                      "Impacto resultado": _FMT_M, "% Ejec.": _FMT_PCT,
                      "% s/Ventas": _FMT_PCT, "Ppto Año": _FMT_M,
                      "% Ppto Año consumido": _FMT_PCT},
            subtotales=subs, resaltar_signo=["Impacto resultado"],
            anchos={"% Ppto Año consumido": 19, "Impacto resultado": 17})
    h.seccion("Márgenes")
    h.tabla(rep["margenes"],
            formatos={"Real": _FMT_PCT, "Ppto": _FMT_PCT, "Δ pp": _FMT_PP},
            resaltar_signo=["Δ pp"])
    h.seccion("Lectura")
    h.texto(rep["conclusiones"])
    h.congelar("B")

    # ── 2. Puente EBIT ────────────────────────────────────────
    h = _Hoja(wb, "2 Puente EBIT", ancho_col_a=42)
    h.titulo("Puente de EBIT — Presupuesto → Real", sub)
    h.texto([
        "El efecto de ventas está valorizado al margen de contribución presupuestado, "
        "y el costo variable se compara contra el que correspondería a las ventas reales. "
        "Así una caída de ventas no aparece como «ahorro» de costo variable.",
        "Signo positivo = suma al EBIT. Signo negativo = lo resta.",
    ])
    h.tabla(rep["puente"], formatos={"Efecto": _FMT_M},
            resaltar_signo=["Efecto"], anchos={"Efecto": 16, "Tipo": 12})
    if abs(rep["descuadre"]) > 1:
        h.texto([f"Descuadre detectado: ${rep['descuadre']:,.0f}"])

    # ── 3. Centros de costo ───────────────────────────────────
    h = _Hoja(wb, "3 Centros de Costo", ancho_col_a=22)
    h.titulo("Gasto por Centro de Costo — Real vs Presupuesto", sub)
    fmt_cc = {"Real YTD": _FMT_M, "Ppto YTD": _FMT_M, "Varianza": _FMT_M,
              "Impacto resultado": _FMT_M, "% Ejec.": _FMT_PCT, "Ppto Año": _FMT_M,
              "% Ppto Año consumido": _FMT_PCT, "Proyección cierre": _FMT_M,
              "Run-rate x12": _FMT_M, "Desv. proyectada": _FMT_M}
    h.tabla(rep["cc"], formatos=fmt_cc, resaltar_signo=["Impacto resultado"],
            anchos={"% Ppto Año consumido": 19, "Proyección cierre": 17,
                    "Impacto resultado": 17, "Desv. proyectada": 16,
                    "Sociedad": 17})
    h.seccion("Apertura por línea del P&L")
    h.tabla(rep["cc_linea"],
            formatos={"Real YTD": _FMT_M, "Ppto YTD": _FMT_M, "Varianza": _FMT_M,
                      "Impacto resultado": _FMT_M, "% Ejec.": _FMT_PCT},
            resaltar_signo=["Impacto resultado"],
            anchos={"Línea": 22, "Impacto resultado": 17})
    h.congelar("C")

    # ── 4. Detalle cuentas ────────────────────────────────────
    h = _Hoja(wb, "4 Detalle Cuentas", ancho_col_a=20)
    h.titulo("Detalle por Cuenta y Centro de Costo", sub)
    h.texto([
        "Ordenado por impacto sobre el resultado. «Tipo de brecha» clasifica el "
        "comportamiento mensual: Recurrente = se repite mes a mes (se anualiza); "
        "Puntual = concentrado en un mes; Desfase de calendario = el acumulado cuadra "
        "pero el gasto cayó en meses distintos; No presupuestado = hay gasto sin presupuesto.",
        "«Meses con desvío desfavorable» indica en qué meses la cuenta se salió del "
        "presupuesto en contra del resultado, y por cuánto: en gastos = se gastó más que lo "
        "presupuestado del mes (signo +); en cuentas de ingreso = se vendió menos (signo −). "
        "Las columnas «Var <mes>» abren la varianza de cada mes (real − presupuesto).",
    ])
    _fmt_det = {"Real YTD": _FMT_M, "Ppto YTD": _FMT_M, "Varianza": _FMT_M,
                "Impacto resultado": _FMT_M, "% Ejec.": _FMT_PCT,
                "Impacto anualizado": _FMT_M, "Ppto Año": _FMT_M,
                "Proyección cierre": _FMT_M, "Meses desviados": _FMT_INT}
    _fmt_det.update({c: _FMT_M for c in rep["cols_var_mes"]})
    _anchos_det = {"Cuenta": 40, "Sociedad": 17, "Línea P&L": 15, "Categoría EERR": 19,
                   "Tipo de brecha": 20, "Impacto anualizado": 18,
                   "Impacto resultado": 17, "Proyección cierre": 17,
                   "Meses desviados": 15, "Meses con desvío desfavorable": 42}
    _anchos_det.update({c: 12 for c in rep["cols_var_mes"]})
    h.tabla(rep["detalle"], formatos=_fmt_det,
            resaltar_signo=["Impacto resultado"], anchos=_anchos_det)
    h.congelar("D")

    # ── 5. Plan de acción ─────────────────────────────────────
    h = _Hoja(wb, "5 Plan de Accion", ancho_col_a=20)
    h.titulo("Plan de Acción — desviaciones que explican la brecha", sub)
    h.texto([
        f"Brecha desfavorable acumulada: ${rep['total_desfavorable']/1e6:,.1f}M. "
        f"Las líneas siguientes concentran el 80% de esa brecha (criterio de Pareto).",
        "Las últimas cuatro columnas van en blanco a propósito: se completan en la "
        "reunión de gerencia con la explicación y el compromiso de cada responsable.",
    ])
    _fmt_ac = {"Real YTD": _FMT_M, "Ppto YTD": _FMT_M, "Varianza": _FMT_M,
               "Impacto resultado": _FMT_M, "% acum. brecha": _FMT_PCT,
               "Impacto anualizado": _FMT_M, "Meses desviados": _FMT_INT}
    _fmt_ac.update({c: _FMT_M for c in rep["cols_var_mes"]})
    _anchos_ac = {"Cuenta": 40, "Sociedad": 17, "Tipo de brecha": 20,
                  "Impacto anualizado": 18,
                  "Impacto resultado": 17, "Meses con desvío desfavorable": 42,
                  "Explicación (qué pasó)": 46, "Acción comprometida": 40,
                  "Responsable": 18, "Fecha": 12}
    _anchos_ac.update({c: 12 for c in rep["cols_var_mes"]})
    h.tabla(rep["accion"], formatos=_fmt_ac,
            resaltar_signo=["Impacto resultado"], anchos=_anchos_ac)
    h.congelar("C")

    # ── 6. Mes a mes ──────────────────────────────────────────
    h = _Hoja(wb, "6 Mes a Mes", ancho_col_a=22)
    h.titulo("Perfil mensual", sub)
    fmt_meses = {m: _FMT_M for m in meta["meses"]}
    fmt_meses["Total YTD"] = _FMT_M
    h.seccion("Gasto por centro de costo")
    h.tabla(rep["mes_cc"], formatos=fmt_meses, anchos={"Concepto": 12})
    h.seccion("Líneas del P&L")
    h.tabla(rep["mes_pl"], formatos=fmt_meses, anchos={"Concepto": 12})
    h.seccion("Cuentas con brecha material — Real / Ppto / Varianza por mes")
    h.texto([
        "Respaldo mes a mes de cada cuenta comentada: muestra exactamente en qué mes "
        "se produjo la desviación y contra qué presupuesto mensual.",
    ])
    h.tabla(rep["mes_cuenta"], formatos=fmt_meses,
            anchos={"Cuenta": 40, "Concepto": 12})
    h.ws.freeze_panes = "D1"

    # ── 7. Proyección ─────────────────────────────────────────
    h = _Hoja(wb, "7 Proyeccion Cierre", ancho_col_a=26)
    h.titulo("Proyección al cierre del año", sub)
    h.texto([
        "«Proyección» = real acumulado + presupuesto de los meses que faltan "
        "(supone que el resto del año se ejecuta según plan).",
        "«Run-rate x12» extrapola el ritmo real actual. En un negocio estacional "
        "el run-rate es referencial: sirve para gastos fijos, no para ventas.",
    ])
    df_pr = rep["proyeccion"]
    subs_pr = [i for i, v in enumerate(df_pr["_subtotal"]) if v]
    h.tabla(df_pr,
            formatos={"Real YTD": _FMT_M, "Ppto restante": _FMT_M,
                      "Proyección (real + ppto restante)": _FMT_M,
                      "Run-rate x12": _FMT_M, "Ppto Año": _FMT_M,
                      "Desv. vs Ppto Año": _FMT_M},
            subtotales=subs_pr, resaltar_signo=["Desv. vs Ppto Año"],
            anchos={"Proyección (real + ppto restante)": 24, "Desv. vs Ppto Año": 18})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── EXPORT HTML ───────────────────────────────────────────────

def _m(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    return f"${v/1e6:,.1f}M"


def _p(v) -> str:
    if v is None or pd.isna(v):
        return "—"
    pct = v * 100
    # Con un decimal bajo el 1% para no mostrar "-0%" en líneas casi nulas
    return f"{pct:,.1f}%" if abs(pct) < 1 else f"{pct:,.0f}%"


def _e(t) -> str:
    return _html.escape(str(t))


def to_html(rep: dict) -> str:
    """Página autocontenida con los insights del reporte (sin dependencias externas)."""
    meta, kpi = rep["meta"], rep["kpi"]

    def kpi_card(label, real, ppto, invertir=False):
        var = real - ppto
        favorable = (var <= 0) if invertir else (var >= 0)
        color = "var(--verde)" if favorable else "var(--rojo)"
        flecha = "▲" if var >= 0 else "▼"
        pct = f"{real/ppto*100:,.0f}%" if ppto else "—"
        return f"""
        <div class="kpi">
          <div class="kpi-l">{_e(label)}</div>
          <div class="kpi-v">{_m(real)}</div>
          <div class="kpi-d" style="color:{color}">{flecha} {_m(abs(var))} vs ppto · {pct}</div>
          <div class="kpi-o">Objetivo {_m(ppto)}</div>
        </div>"""

    # Puente en SVG
    puente = rep["puente"]
    efectos = puente[puente["Tipo"] == "efecto"]
    ini = float(puente.iloc[0]["Efecto"])
    fin = float(puente.iloc[-1]["Efecto"])
    barras = [("EBIT Ppto", ini, "total")]
    for _, r in efectos.iterrows():
        barras.append((str(r["Concepto"]), float(r["Efecto"]), "efecto"))
    barras.append(("EBIT Real", fin, "total"))

    acum, puntos = ini, []
    for nombre, val, tipo in barras:
        if tipo == "total":
            # Las barras de total arrancan del cero; el presupuesto de estos
            # meses puede ser negativo (baja temporada), así que el rango tiene
            # que admitir barras hacia abajo.
            lo, hi = min(0.0, val), max(0.0, val)
        else:
            base = acum
            acum += val
            lo, hi = min(base, acum), max(base, acum)
        puntos.append((nombre, lo, hi, val, tipo))

    vals = [p[2] for p in puntos] + [p[1] for p in puntos] + [0.0]
    vmax, vmin = max(vals), min(vals)
    rango = (vmax - vmin) or 1

    ancho_barra, sep = 62, 20
    w = max(760, len(puntos) * (ancho_barra + sep) + 60)
    alto_plot, y_base, hgt = 190, 235, 320   # y_base = piso del área de barras

    def _y(v: float) -> float:
        return y_base - (v - vmin) / rango * alto_plot

    y_cero = _y(0.0)
    svg = []
    for i, (nombre, lo, hi, val, tipo) in enumerate(puntos):
        x = 40 + i * (ancho_barra + sep)
        y_top, y_bot = _y(hi), _y(lo)
        color = "#2D0050" if tipo == "total" else ("#0F6E56" if val >= 0 else "#C4007A")
        alto = max(y_bot - y_top, 2)
        svg.append(f'<rect x="{x}" y="{y_top:.1f}" width="{ancho_barra}" height="{alto:.1f}" '
                   f'fill="{color}" rx="3"/>')
        # En los totales el signo es parte de la cifra (el EBIT presupuestado de
        # estos meses puede ser negativo); en los efectos indica si suma o resta.
        prefijo = ("−" if val < 0 else "") if tipo == "total" else ("+" if val >= 0 else "−")
        svg.append(f'<text x="{x + ancho_barra/2:.0f}" y="{y_top - 7:.1f}" text-anchor="middle" '
                   f'font-size="10.5" font-weight="700" fill="{color}">'
                   f'{prefijo}{abs(val)/1e6:,.1f}M</text>')
        etiqueta = nombre if len(nombre) <= 24 else nombre[:23] + "…"
        y_lbl = y_base + 16
        svg.append(f'<text x="{x + ancho_barra/2:.0f}" y="{y_lbl}" text-anchor="end" '
                   f'font-size="9.5" fill="#666" '
                   f'transform="rotate(-35 {x + ancho_barra/2:.0f} {y_lbl})">'
                   f'{_e(etiqueta)}</text>')
    svg_puente = (f'<svg viewBox="0 0 {w} {hgt}" width="100%" preserveAspectRatio="xMinYMin meet" '
                  f'role="img" aria-label="Puente de EBIT">'
                  f'<line x1="30" y1="{y_cero:.1f}" x2="{w-10}" y2="{y_cero:.1f}" '
                  f'stroke="#D9C7E6" stroke-dasharray="3 3"/>'
                  f'<text x="24" y="{y_cero + 3:.1f}" text-anchor="end" font-size="9" '
                  f'fill="#A9A9B8">0</text>'
                  + "".join(svg) + "</svg>")

    # Barras por centro de costo
    df_cc = rep["cc"]
    filas_cc = []
    if not df_cc.empty:
        tope = max(float(df_cc[["Real YTD", "Ppto YTD"]].max().max()), 1)
        for _, r in df_cc.iterrows():
            pr = float(r["Real YTD"]) / tope * 100
            pp = float(r["Ppto YTD"]) / tope * 100
            imp = float(r["Impacto resultado"])
            color = "var(--verde)" if imp >= 0 else "var(--rojo)"
            filas_cc.append(f"""
            <div class="ccrow">
              <div class="ccname">{_e(r['Centro de costo'])}
                <span class="ccsub">{_p(r['% Ppto Año consumido'])} del ppto anual ·
                  {_e(r['Sociedad'])}</span></div>
              <div class="ccbars">
                <div class="bar real" style="width:{pr:.1f}%"></div>
                <div class="tick" style="left:{pp:.1f}%"></div>
              </div>
              <div class="ccval">{_m(r['Real YTD'])}<span class="ccppto"> / {_m(r['Ppto YTD'])}</span></div>
              <div class="ccimp" style="color:{color}">{'+' if imp >= 0 else '−'}{_m(abs(imp))}</div>
            </div>""")

    # Tabla de acción
    df_ac = rep["accion"]
    filas_ac = []
    piso_mes = meta["umbral"] / max(meta["n_meses"], 1)

    def _strip_meses(real_mes, ppto_mes) -> str:
        """Franja mes a mes: marca en qué meses la cuenta se pasó del presupuesto."""
        chips = []
        for i, etiqueta in enumerate(meta["meses"]):
            rv = float(real_mes[i])
            pv = float(ppto_mes[i])
            dif = rv - pv
            if dif > piso_mes:
                clase = "up"
            elif dif < -piso_mes:
                clase = "down"
            else:
                clase = "flat"
            tip = (f"{etiqueta}: real {_m(rv)} · ppto {_m(pv)} · "
                   f"{'+' if dif >= 0 else '−'}{_m(abs(dif))}")
            chips.append(f'<span class="mchip {clase}" title="{_e(tip)}">{_e(etiqueta)}</span>')
        return f'<div class="strip">{"".join(chips)}</div>'

    if not df_ac.empty:
        for _, r in df_ac.iterrows():
            imp = float(r["Impacto resultado"])
            tipo = str(r["Tipo de brecha"])
            clase = {"Recurrente": "tag-rec", "Puntual": "tag-pun",
                     "No presupuestado": "tag-nop",
                     "Desfase de calendario": "tag-des"}.get(tipo, "tag-mix")
            anual = float(r["Impacto anualizado"]) if pd.notna(r["Impacto anualizado"]) else 0.0
            filas_ac.append(f"""
            <tr>
              <td><b>{_e(r['Cuenta'])}</b>
                  <div class="sub">{_e(r['Centro de costo'])} ·
                    <span class="soc">{_e(r['Sociedad'])}</span></div>
                  {_strip_meses(r['_real_mes'], r['_ppto_mes'])}
                  <div class="sub">Desfavorable: {_e(r['Meses con desvío desfavorable'])}</div></td>
              <td class="num">{_m(r['Real YTD'])}</td>
              <td class="num">{_m(r['Ppto YTD'])}</td>
              <td class="num neg">{_m(imp)}</td>
              <td class="num">{_p(r['% acum. brecha'])}</td>
              <td><span class="tag {clase}">{_e(tipo)}</span>
                  <div class="sub">{_plural(int(r['Meses desviados']), 'mes', 'meses')} ·
                    pico {_e(r['Mes pico'])}</div></td>
              <td class="num">{_m(anual) if anual else '—'}</td>
            </tr>""")

    # P&L
    filas_pl = []
    for _, r in rep["pl"].iterrows():
        sub = bool(r["_subtotal"])
        var = float(r["Varianza"])
        imp = float(r["Impacto resultado"])
        filas_pl.append(f"""
        <tr class="{'sub' if sub else ''}">
          <td>{_e(r['Línea'])}</td>
          <td class="num">{_m(r['Real YTD'])}</td>
          <td class="num">{_m(r['Ppto YTD'])}</td>
          <td class="num">{_m(var)}</td>
          <td class="num {'pos' if imp >= 0 else 'neg'}">{'+' if imp >= 0 else '−'}{_m(abs(imp))}</td>
          <td class="num">{_p(r['% Ejec.'])}</td>
          <td class="num">{_p(r['% Ppto Año consumido'])}</td>
        </tr>""")

    def _clase_pp(delta: float) -> str:
        # Bajo una décima de punto no hay noticia: se muestra neutro
        if abs(delta) < 0.05:
            return ""
        return "pos" if delta > 0 else "neg"

    filas_mg = "".join(
        f"<tr><td>{_e(r['Margen'])}</td><td class='num'>{_p(r['Real'])}</td>"
        f"<td class='num'>{_p(r['Ppto'])}</td>"
        f"<td class='num {_clase_pp(r['Δ pp'])}'>{r['Δ pp']:+.1f} pp</td></tr>"
        for _, r in rep["margenes"].iterrows())

    alertas = "".join(f"<li>{_e(a)}</li>" for a in rep["alertas"])
    escalas = "".join(f"<li>{_e(x)}</li>" for x in rep["escalas"])
    conclusiones = "".join(f"<li>{_e(c)}</li>" for c in rep["conclusiones"])
    bases = "".join(f"<li>{_e(b)}</li>" for b in rep["bases"])

    return f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Reporte de Gerencia · Kreems · {_e(meta['periodo_lbl'])}</title>
<style>
  :root {{
    --morado:#2D0050; --morado2:#6B2C91; --fucsia:#C4007A;
    --verde:#0F6E56; --rojo:#CC0000; --borde:#EDE4F3; --gris:#7A7A8C;
  }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; background:#F7F5FA; color:#22222E;
         font-family:'Inter',-apple-system,'Segoe UI',Roboto,sans-serif; font-size:14px; }}
  .wrap {{ max-width:1180px; margin:0 auto; padding:0 22px 64px; }}
  header {{ background:var(--morado); color:#fff; padding:26px 0 22px; margin-bottom:26px; }}
  header .wrap {{ padding-bottom:0; }}
  h1 {{ margin:0 0 6px; font-size:25px; font-weight:800; letter-spacing:-0.3px; }}
  .hsub {{ opacity:.72; font-size:13.5px; }}
  .hmix {{ opacity:.55; font-size:11.5px; margin-top:3px; }}
  h2 {{ font-size:17px; color:var(--morado); margin:34px 0 12px;
        border-bottom:2px solid var(--borde); padding-bottom:7px; }}
  .kpis {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(210px,1fr)); gap:14px; }}
  .kpi {{ background:#fff; border:1px solid var(--borde); border-radius:12px; padding:15px 17px; }}
  .kpi-l {{ font-size:11px; color:var(--gris); text-transform:uppercase;
            letter-spacing:.5px; font-weight:600; }}
  .kpi-v {{ font-size:26px; font-weight:800; color:var(--morado); margin:5px 0 2px; }}
  .kpi-d {{ font-size:12.5px; font-weight:600; }}
  .kpi-o {{ font-size:11px; color:#A9A9B8; margin-top:2px; }}
  .card {{ background:#fff; border:1px solid var(--borde); border-radius:12px;
           padding:18px 20px; margin-top:12px; }}
  table {{ width:100%; border-collapse:collapse; font-size:13px; }}
  th {{ background:var(--morado2); color:#fff; text-align:right; padding:9px 10px;
        font-size:11.5px; font-weight:600; }}
  th:first-child {{ text-align:left; border-radius:6px 0 0 0; }}
  th:last-child {{ border-radius:0 6px 0 0; }}
  td {{ padding:8px 10px; border-bottom:1px solid #F2ECF6; }}
  td.num {{ text-align:right; font-variant-numeric:tabular-nums; }}
  tr.sub td {{ background:#FAF5FD; font-weight:700; color:var(--morado); }}
  .pos {{ color:var(--verde); font-weight:600; }}
  .neg {{ color:var(--rojo); font-weight:600; }}
  .sub {{ font-size:11px; color:var(--gris); font-weight:400; }}
  .soc {{ font-size:10px; background:#F1EDF6; color:var(--morado2);
          padding:1px 6px; border-radius:4px; font-weight:600; white-space:nowrap; }}
  ul.ins {{ margin:0; padding-left:20px; }}
  ul.ins li {{ margin-bottom:9px; line-height:1.6; }}
  .alerta {{ background:#FFF6F6; border:1px solid #F3D0D0; border-left:4px solid var(--rojo);
             border-radius:9px; padding:14px 18px; }}
  .alerta ul {{ margin:0; padding-left:19px; }}
  .alerta li {{ margin-bottom:7px; line-height:1.55; }}
  .ccrow {{ display:grid; grid-template-columns:190px 1fr 165px 95px;
            align-items:center; gap:12px; padding:9px 0; border-bottom:1px solid #F2ECF6; }}
  .ccname {{ font-weight:600; font-size:13px; }}
  .ccsub {{ display:block; font-size:11px; color:var(--gris); font-weight:400; }}
  .ccbars {{ position:relative; height:16px; background:#F4EFF8; border-radius:8px; }}
  .bar.real {{ position:absolute; left:0; top:0; height:16px; background:var(--fucsia);
               border-radius:8px; opacity:.85; }}
  .tick {{ position:absolute; top:-3px; width:2.5px; height:22px; background:var(--morado); }}
  .ccval {{ text-align:right; font-size:13px; font-weight:600;
            font-variant-numeric:tabular-nums; }}
  .ccppto {{ color:var(--gris); font-weight:400; }}
  .ccimp {{ text-align:right; font-weight:700; font-variant-numeric:tabular-nums; }}
  .tag {{ display:inline-block; padding:2px 9px; border-radius:20px;
          font-size:10.5px; font-weight:700; white-space:nowrap; }}
  .tag-rec {{ background:#FDE8E8; color:#B01919; }}
  .tag-pun {{ background:#FFF2DC; color:#96620A; }}
  .tag-nop {{ background:#F0E6FA; color:#5B2A87; }}
  .tag-des {{ background:#E6F1FB; color:#1B5E96; }}
  .tag-mix {{ background:#EFEFF3; color:#55555F; }}
  .strip {{ display:flex; gap:3px; margin:6px 0 4px; flex-wrap:wrap; }}
  .mchip {{ font-size:9.5px; font-weight:700; padding:1.5px 6px; border-radius:4px;
            background:#F1EFF4; color:#9A9AA8; cursor:default; }}
  .mchip.up {{ background:#FDE8E8; color:#B01919; }}
  .mchip.down {{ background:#E4F3EE; color:#0F6E56; }}
  .leyenda {{ font-size:11.5px; color:var(--gris); margin-top:10px; line-height:1.6; }}
  footer {{ margin-top:40px; font-size:11.5px; color:var(--gris); line-height:1.7; }}
  .scroll {{ overflow-x:auto; }}
  @media print {{ body {{ background:#fff; }} .card {{ break-inside:avoid; }} }}
</style></head><body>
<header><div class="wrap">
  <h1>Reporte de Gerencia — Real vs Presupuesto</h1>
  <div class="hsub">Kreems · {_e(meta['sociedad'])} · {_e(meta['periodo_lbl'])}
   · {meta['n_meses']} de 12 meses · generado {_e(meta['generado'])}</div>
  <div class="hmix">Facturación del periodo: {_e(meta['mix_sociedad'])}</div>
</div></header>
<div class="wrap">

  <div class="kpis">
    {kpi_card("Ventas", kpi['ventas_r'], kpi['ventas_p'])}
    {kpi_card("Utilidad Bruta", kpi['ub_r'], kpi['ub_p'])}
    {kpi_card("EBIT", kpi['ebit_r'], kpi['ebit_p'])}
    {kpi_card("Gasto total", kpi['gasto_r'], kpi['gasto_p'], invertir=True)}
  </div>

  {"<h2>Advertencias sobre la base de comparación</h2><div class='alerta'><ul>" + alertas + "</ul></div>" if alertas else ""}

  <h2>Conclusiones</h2>
  <div class="card"><ul class="ins">{conclusiones}</ul></div>

  <h2>Estado de Resultados acumulado</h2>
  <div class="card scroll">
    <table>
      <thead><tr><th>Línea</th><th>Real YTD</th><th>Ppto YTD</th><th>Varianza</th>
        <th>Impacto resultado</th><th>% Ejec.</th><th>% Ppto Año</th></tr></thead>
      <tbody>{''.join(filas_pl)}</tbody>
    </table>
    <div class="leyenda">«Ppto YTD» es el presupuesto de los mismos meses acumulados,
      no el anual. «Impacto resultado» ya viene con el signo del efecto sobre el EBIT:
      en gastos, gastar de más resta.</div>
  </div>
  <div class="card scroll">
    <table>
      <thead><tr><th>Margen</th><th>Real</th><th>Ppto</th><th>Δ</th></tr></thead>
      <tbody>{filas_mg}</tbody>
    </table>
  </div>

  <h2>Puente de EBIT — de dónde sale la diferencia</h2>
  <div class="card scroll">
    {svg_puente}
    <div class="leyenda">El efecto de ventas está valorizado al margen de contribución
      presupuestado y el costo variable se mide contra el que correspondería a las ventas
      reales, para que una caída de ventas no aparezca como ahorro de costo.</div>
  </div>

  <h2>Gasto por centro de costo</h2>
  <div class="card">
    {''.join(filas_cc) if filas_cc else '<div class="leyenda">Sin datos.</div>'}
    <div class="leyenda">Barra fucsia = gasto real acumulado · marca morada = presupuesto
      de los mismos meses · última columna = impacto sobre el resultado.</div>
  </div>

  <h2>Desviaciones que explican la brecha</h2>
  <div class="card scroll">
    <table>
      <thead><tr><th>Cuenta / centro de costo</th><th>Real YTD</th><th>Ppto YTD</th>
        <th>Impacto</th><th>% acum. brecha</th><th>Tipo de brecha</th>
        <th>Si se mantiene 12m</th></tr></thead>
      <tbody>{''.join(filas_ac) if filas_ac else '<tr><td colspan="7">Sin desviaciones sobre el umbral de materialidad.</td></tr>'}</tbody>
    </table>
    <div class="leyenda">
      La franja de meses bajo cada cuenta marca en <span class="mchip up">rojo</span> los
      meses en que se pasó del presupuesto, en <span class="mchip down">verde</span> los que
      quedaron bajo y en <span class="mchip">gris</span> los que fueron en línea (pasa el
      cursor por encima para ver real, presupuesto y diferencia del mes).<br>
      <b>Recurrente</b>: se repite mes a mes, es estructural y se puede anualizar ·
      <b>Puntual</b>: concentrado en un mes ·
      <b>Desfase de calendario</b>: el acumulado cuadra, el gasto cayó en otros meses ·
      <b>No presupuestado</b>: hay gasto sin presupuesto asignado.
    </div>
  </div>

  <footer>
    <b>Cómo leer las cifras</b>
    <ul>{escalas}</ul>
    <b>Bases y criterios</b>
    <ul>{bases}</ul>
  </footer>
</div></body></html>"""
