"""
ETL functions — ACUÑA, Gran Natural, Presupuesto, CV Real Sync
Adaptados para Streamlit: reciben bytes del file_uploader,
usan get_engine() de db.py (Supabase via st.secrets).
"""
import io
import re
import unicodedata
import pandas as pd
from datetime import date
from sqlalchemy import text
from .db import get_engine

# ── Mapeos CC ─────────────────────────────────────────────────
MAPA_CC_ACUNA = {
    "NINGUNO":          "CC-00",
    "ADMINISTRACION":   "CC-01",
    "COSTO FABRICA":    "CC-04",
    "DISTRIBUCION":     "CC-03",
    "VENTAS":           "CC-02",
    "GERENCIA":         "CC-01",
    "MAQUINA COMODATO": "CC-03",
}

MAPA_CC_GN = {
    "Ninguno":        "CC-00",
    "Administracion": "CC-01",
    "Comercial":      "CC-02",
    "Distribucion":   "CC-03",
    "Produccion":     "CC-04",
}

HOJAS_CC_PPTO = {
    "ADMINISTRACIÓN": "CC-01",
    "PRODUCCIÓN":     "CC-04",
    "COMERCIAL":      "CC-02",
    "DISTRIBUCIÓN":   "CC-03",
}

MESES_PPTO = {
    "Ene": "01", "Enero": "01", "Feb": "02", "Febrero": "02",
    "Mar": "03", "Marzo": "03", "Abr": "04", "Abril": "04",
    "May": "05", "Mayo": "05", "Jun": "06", "Junio": "06",
    "Jul": "07", "Julio": "07", "Ago": "08", "Agosto": "08",
    "Sep": "09", "Septiembre": "09", "Oct": "10", "Octubre": "10",
    "Nov": "11", "Noviembre": "11", "Dic": "12", "Diciembre": "12",
}

FILAS_IGNORAR = {
    "total ingresos", "total gastos", "gastos", "ingresos",
    "cuenta", "resultado del ejercicio", "total"
}


def _log(lines: list, msg: str):
    lines.append(f"  {msg}")


def _extraer_periodo_obuma(ws) -> str:
    """Busca 'Desde el DD-MM-YYYY' en las primeras 5 filas del workbook."""
    for row in ws.iter_rows(max_row=5, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                match = re.search(r"Desde el[\s\xa0]+(\d{2})-(\d{2})-(\d{4})", cell)
                if match:
                    _, mes, anio = match.groups()
                    return f"{anio}-{mes}"
    raise ValueError("No se encontró el periodo en el encabezado. Revisa el formato del archivo.")


def _registrar_auditoria(engine, tabla, periodo, n, observaciones):
    try:
        with engine.begin() as conn:
            conn.execute(text("""
                INSERT INTO audit.log_carga
                    (tabla_destino, archivo_origen, periodo, registros_cargados, estado, observaciones)
                VALUES (:t, 'webapp_upload', :p, :n, 'OK', :obs)
            """), {"t": tabla, "p": periodo, "n": n, "obs": observaciones})
    except Exception:
        pass  # El log no debe romper la carga principal


# ═══════════════════════════════════════════════════════════════
# ETL ACUÑA
# ═══════════════════════════════════════════════════════════════

def run_etl_acuna(file_bytes: bytes) -> dict:
    """
    Procesa Excel Obuma ACUÑA → marts.fact_real (sociedad = 'ACUÑA').
    Usa dim_homologacion para mapear cuentas ACUÑA → plan cuentas GN.
    """
    import openpyxl
    logs = []
    engine = get_engine()

    try:
        _log(logs, "Abriendo archivo Excel ACUÑA...")
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        periodo = _extraer_periodo_obuma(ws)
        _log(logs, f"Periodo detectado: {periodo}")

        # Cargar tabla de homologación
        with engine.connect() as conn:
            rows = conn.execute(text("""
                SELECT codigo_acuna, codigo_gn
                FROM master.dim_homologacion
                WHERE sociedad_origen = 'ACUÑA' AND activo = TRUE
            """)).fetchall()
        homologacion = {r[0]: (r[1] if r[1] else r[0]) for r in rows}
        _log(logs, f"Homologación cargada: {len(homologacion)} cuentas mapeadas")

        # Detectar columnas CC desde encabezado (fila 6, índice 5)
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 7:
            raise ValueError("El archivo no tiene suficientes filas. ¿Es el archivo correcto?")

        headers = [str(h).strip().upper() if h else "" for h in all_rows[5]]
        MAPA_UPPER = {k.upper(): v for k, v in MAPA_CC_ACUNA.items()}
        cc_indices = {i: MAPA_UPPER[h] for i, h in enumerate(headers) if h in MAPA_UPPER}

        if not cc_indices:
            raise ValueError(f"No se encontraron columnas CC válidas. Encabezados: {headers}")
        _log(logs, f"Columnas CC detectadas: {list(set(cc_indices.values()))}")

        # Transformar filas
        registros = []
        for row in all_rows[6:]:
            celda = row[0]
            if not celda:
                continue
            celda_str = str(celda).strip()
            if celda_str.lower() in FILAS_IGNORAR or celda_str.startswith("Total"):
                continue
            codigo_raw = celda_str.split(" ", 1)[0].strip()
            if "." not in codigo_raw:
                continue
            if codigo_raw not in homologacion:
                continue
            codigo_gn = homologacion[codigo_raw]

            for idx, codigo_cc in cc_indices.items():
                if idx >= len(row) or row[idx] is None:
                    continue
                try:
                    monto = float(row[idx])
                except (TypeError, ValueError):
                    continue
                if monto == 0:
                    continue
                registros.append({
                    "fecha":          pd.to_datetime(f"{periodo}-01"),
                    "codigo_cuenta":  codigo_gn,
                    "codigo_cc":      codigo_cc,
                    "valor":          monto,
                    "periodo":        periodo,
                    "fuente":         "OBUMA",
                    "archivo_origen": "webapp_upload",
                    "sociedad":       "ACUÑA",
                })

        df = pd.DataFrame(registros)
        _log(logs, f"Registros procesados: {len(df)}")
        if df.empty:
            raise ValueError("No se encontraron registros válidos. Verifica que las cuentas estén en la tabla de homologación.")

        # Excluir cuenta 3.1.01.001: se gestiona exclusivamente via staging.cv_real_manual
        antes = len(df)
        df = df[df["codigo_cuenta"] != CODIGO_CUENTA_CV].copy()
        if len(df) < antes:
            _log(logs, f"Cuenta {CODIGO_CUENTA_CV} excluida del ETL (se gestiona via CV staging)")

        # Resumen por CC
        for cc, val in df.groupby("codigo_cc")["valor"].sum().items():
            _log(logs, f"  {cc}: ${val:,.0f}")

        # Cargar a BD
        with engine.begin() as conn:
            # Excluir fuente='CV_MANUAL' para preservar el costo variable ingresado manualmente
            r = conn.execute(text("""
                DELETE FROM marts.fact_real
                WHERE periodo = :p AND sociedad = 'ACUÑA' AND fuente <> 'CV_MANUAL'
            """), {"p": periodo})
            _log(logs, f"Registros anteriores eliminados: {r.rowcount} (CV Manual preservado)")

            df.to_sql("fact_real", con=conn, schema="marts", if_exists="append", index=False)

            conn.execute(text("""
                UPDATE marts.fact_real
                SET fecha_id = TO_CHAR(fecha, 'YYYYMMDD')::INT
                WHERE fecha_id IS NULL AND periodo = :p AND sociedad = 'ACUÑA'
            """), {"p": periodo})

        _log(logs, f"✓ {len(df)} registros cargados en marts.fact_real")
        _registrar_auditoria(engine, "marts.fact_real", periodo, len(df), "ETL ACUÑA via webapp")

        return {"ok": True, "periodo": periodo, "n_registros": len(df), "logs": logs, "error": None}

    except Exception as e:
        _log(logs, f"✗ Error: {e}")
        return {"ok": False, "periodo": None, "n_registros": 0, "logs": logs, "error": str(e)}


# ═══════════════════════════════════════════════════════════════
# ETL GRAN NATURAL
# ═══════════════════════════════════════════════════════════════

def run_etl_gn(file_bytes: bytes) -> dict:
    """
    Procesa Excel Obuma Gran Natural → marts.fact_real (sociedad = 'GRAN_NATURAL').
    """
    import openpyxl
    logs = []
    engine = get_engine()

    try:
        _log(logs, "Abriendo archivo Excel Gran Natural...")
        # Extraer periodo con openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        periodo = _extraer_periodo_obuma(ws)
        _log(logs, f"Periodo detectado: {periodo}")

        # Leer con pandas para el melt
        df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
        _log(logs, f"Filas leídas: {len(df_raw)}")

        # El archivo GN tiene 7 columnas (Cuenta + 5 CC + Total)
        n_cols = len(df_raw.columns)
        if n_cols < 6:
            raise ValueError(f"El archivo tiene solo {n_cols} columnas. Se esperan al menos 6.")

        # Asignar nombres de columna estándar
        col_names = ["Cuenta", "Ninguno", "Administracion", "Comercial", "Distribucion", "Produccion"]
        if n_cols >= 7:
            col_names.append("Total")
        df_raw.columns = col_names + list(range(n_cols - len(col_names)))

        # Filtrar solo filas con código de cuenta (contienen ".")
        df = df_raw[
            df_raw["Cuenta"].notna() &
            df_raw["Cuenta"].astype(str).str.contains(r"\.", regex=True) &
            ~df_raw["Cuenta"].astype(str).str.startswith("Desde") &
            ~df_raw["Cuenta"].astype(str).str.startswith("Centro")
        ].copy()

        if df.empty:
            raise ValueError("No se encontraron filas con código de cuenta. ¿Es el archivo GN correcto?")

        df["codigo_cuenta"] = df["Cuenta"].astype(str).str.split(" ", n=1).str[0].str.strip()

        cols_cc = ["Ninguno", "Administracion", "Comercial", "Distribucion", "Produccion"]
        df_long = df.melt(
            id_vars=["codigo_cuenta"],
            value_vars=cols_cc,
            var_name="nombre_cc_raw",
            value_name="valor"
        )
        df_long["valor"] = pd.to_numeric(df_long["valor"], errors="coerce")
        df_long = df_long[df_long["valor"].notna() & (df_long["valor"] != 0)].copy()
        df_long["codigo_cc"]      = df_long["nombre_cc_raw"].map(MAPA_CC_GN)
        df_long["periodo"]        = periodo
        df_long["fecha"]          = pd.to_datetime(f"{periodo}-01")
        df_long["fuente"]         = "OBUMA"
        df_long["archivo_origen"] = "webapp_upload"
        df_long["sociedad"]       = "GRAN_NATURAL"

        df_final = df_long[["fecha", "codigo_cuenta", "codigo_cc",
                             "valor", "periodo", "fuente", "archivo_origen", "sociedad"]]

        _log(logs, f"Registros procesados: {len(df_final)}")
        if df_final.empty:
            raise ValueError("No se encontraron registros válidos después del filtrado.")

        # Excluir cuenta 3.1.01.001: se gestiona exclusivamente via staging.cv_real_manual
        antes_gn = len(df_final)
        df_final = df_final[df_final["codigo_cuenta"] != CODIGO_CUENTA_CV].copy()
        if len(df_final) < antes_gn:
            _log(logs, f"Cuenta {CODIGO_CUENTA_CV} excluida del ETL (se gestiona via CV staging)")

        for cc, val in df_final.groupby("codigo_cc")["valor"].sum().items():
            _log(logs, f"  {cc}: ${val:,.0f}")

        with engine.begin() as conn:
            # Excluir fuente='CV_MANUAL' para preservar el costo variable ingresado manualmente
            r = conn.execute(text("""
                DELETE FROM marts.fact_real
                WHERE periodo = :p AND sociedad = 'GRAN_NATURAL' AND fuente <> 'CV_MANUAL'
            """), {"p": periodo})
            _log(logs, f"Registros anteriores eliminados: {r.rowcount} (CV Manual preservado)")

            df_final.to_sql("fact_real", con=conn, schema="marts", if_exists="append", index=False)

            conn.execute(text("""
                UPDATE marts.fact_real
                SET fecha_id = TO_CHAR(fecha, 'YYYYMMDD')::INT
                WHERE fecha_id IS NULL AND periodo = :p AND sociedad = 'GRAN_NATURAL'
            """), {"p": periodo})

        _log(logs, f"✓ {len(df_final)} registros cargados en marts.fact_real")
        _registrar_auditoria(engine, "marts.fact_real", periodo, len(df_final), "ETL GRAN NATURAL via webapp")

        return {"ok": True, "periodo": periodo, "n_registros": len(df_final), "logs": logs, "error": None}

    except Exception as e:
        _log(logs, f"✗ Error: {e}")
        return {"ok": False, "periodo": None, "n_registros": 0, "logs": logs, "error": str(e)}


# ═══════════════════════════════════════════════════════════════
# ETL PRESUPUESTO
# ═══════════════════════════════════════════════════════════════

def run_etl_presupuesto(file_bytes: bytes, anio: str | None = None) -> dict:
    """
    Procesa Presupuesto_Maestro.xlsx → marts.fact_presupuesto.
    Reemplaza todo el presupuesto del año indicado (por defecto, el año actual).
    """
    if anio is None:
        anio = str(date.today().year)
    logs = []
    engine = get_engine()
    MESES = {k: f"{anio}-{v}" for k, v in MESES_PPTO.items()}

    def leer_hoja_cc(hoja, codigo_cc):
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=hoja, header=None)
        header_row = next(
            (i for i, v in enumerate(df.iloc[:, 0]) if str(v).strip() == "CÓDIGO"), None
        )
        if header_row is None:
            raise ValueError(f"No se encontró 'CÓDIGO' en hoja '{hoja}'")
        df.columns = df.iloc[header_row].astype(str).str.strip()
        df = df.iloc[header_row + 1:].reset_index(drop=True)
        cols = list(df.columns)
        cols[0] = "codigo_cuenta"
        df.columns = cols
        df["codigo_cuenta"] = df["codigo_cuenta"].astype(str).str.strip()
        df["codigo_cuenta"] = df["codigo_cuenta"].where(
            df["codigo_cuenta"].str.match(r'^\d+\.\d+\.\d+\.\d+$')
        ).ffill()
        df = df[df["codigo_cuenta"].str.match(r'^\d+\.\d+\.\d+\.\d+$', na=False)].copy()
        cols_mes = [c for c in df.columns if c in MESES]
        if not cols_mes:
            raise ValueError(f"No se encontraron columnas de meses en hoja '{hoja}'")
        for col in cols_mes:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        df_agg = df.groupby("codigo_cuenta")[cols_mes].sum().reset_index()
        df_long = df_agg.melt(id_vars=["codigo_cuenta"], value_vars=cols_mes,
                              var_name="mes", value_name="valor")
        df_long["periodo"]        = df_long["mes"].map(MESES)
        df_long["fecha"]          = pd.to_datetime(df_long["periodo"] + "-01")
        df_long["codigo_cc"]      = codigo_cc
        df_long["fuente"]         = "PRESUPUESTO"
        df_long["archivo_origen"] = "webapp_upload"
        df_long = df_long[df_long["valor"] != 0]
        return df_long[["fecha", "codigo_cuenta", "codigo_cc", "valor",
                         "periodo", "fuente", "archivo_origen"]]

    def leer_consolidado():
        df = pd.read_excel(io.BytesIO(file_bytes),
                           sheet_name="Consolidado_Automatico", header=None)
        registros = []

        # Ventas (4.1.01.001): fila 2 = meses, fila 3 = valores
        meses_v   = [str(v).strip() for v in df.iloc[2, 0:12]]
        valores_v = df.iloc[3, 0:12].tolist()
        for mes, val in zip(meses_v, valores_v):
            if mes in MESES and pd.notna(val) and float(val) != 0:
                registros.append({
                    "fecha": pd.to_datetime(MESES[mes] + "-01"),
                    "codigo_cuenta": "4.1.01.001", "codigo_cc": "CC-00",
                    "valor": float(val), "periodo": MESES[mes],
                    "fuente": "PRESUPUESTO", "archivo_origen": "webapp_upload",
                })

        # CV (3.1.01.001): filas 6+7 y 11+12 (dos líneas de CV)
        for fila_mes, fila_val in [(6, 7), (11, 12)]:
            meses_cv  = [str(v).strip() for v in df.iloc[fila_mes, 0:12]]
            valores_cv = df.iloc[fila_val, 0:12].tolist()
            for mes, val in zip(meses_cv, valores_cv):
                if mes in MESES and pd.notna(val) and float(val) != 0:
                    existing = next(
                        (r for r in registros
                         if r["periodo"] == MESES[mes] and r["codigo_cuenta"] == "3.1.01.001"),
                        None
                    )
                    if existing:
                        existing["valor"] += float(val)
                    else:
                        registros.append({
                            "fecha": pd.to_datetime(MESES[mes] + "-01"),
                            "codigo_cuenta": "3.1.01.001", "codigo_cc": "CC-00",
                            "valor": float(val), "periodo": MESES[mes],
                            "fuente": "PRESUPUESTO", "archivo_origen": "webapp_upload",
                        })
        return pd.DataFrame(registros)

    try:
        _log(logs, f"Procesando Presupuesto Maestro {anio}...")
        frames = []

        for hoja, codigo_cc in HOJAS_CC_PPTO.items():
            _log(logs, f"  Leyendo hoja {hoja} ({codigo_cc})...")
            df_cc = leer_hoja_cc(hoja, codigo_cc)
            _log(logs, f"    → {len(df_cc)} registros")
            frames.append(df_cc)

        _log(logs, "  Leyendo Consolidado_Automatico (Ventas + CV)...")
        df_consol = leer_consolidado()
        _log(logs, f"    → {len(df_consol)} registros")
        if not df_consol.empty:
            frames.append(df_consol)

        df_final = pd.concat(frames, ignore_index=True)
        _log(logs, f"Total registros a cargar: {len(df_final)}")

        # Resumen por CC (total anual)
        for cc, val in df_final.groupby("codigo_cc")["valor"].sum().items():
            _log(logs, f"  {cc}: ${val:,.0f}")

        with engine.begin() as conn:
            r = conn.execute(text(
                "DELETE FROM marts.fact_presupuesto WHERE periodo LIKE :anio"
            ), {"anio": f"{anio}-%"})
            _log(logs, f"Registros anteriores eliminados: {r.rowcount}")

            df_final.to_sql("fact_presupuesto", con=conn, schema="marts",
                            if_exists="append", index=False, method="multi")

            conn.execute(text("""
                UPDATE marts.fact_presupuesto
                SET fecha_id = TO_CHAR(fecha, 'YYYYMMDD')::INT
                WHERE fecha_id IS NULL AND periodo LIKE :anio
            """), {"anio": f"{anio}-%"})

        _log(logs, f"✓ {len(df_final)} registros cargados en marts.fact_presupuesto")
        _registrar_auditoria(engine, "marts.fact_presupuesto",
                             f"{anio}-AN", len(df_final), "ETL Presupuesto via webapp")

        return {"ok": True, "periodo": f"{anio} (anual)", "n_registros": len(df_final),
                "logs": logs, "error": None}

    except Exception as e:
        _log(logs, f"✗ Error: {e}")
        return {"ok": False, "periodo": None, "n_registros": 0, "logs": logs, "error": str(e)}


# ═══════════════════════════════════════════════════════════════
# CV REAL SYNC
# ═══════════════════════════════════════════════════════════════

CODIGO_CUENTA_CV = "3.1.01.001"
CODIGO_CC_CV     = "CC-00"

PERIODOS_2026 = {
    f"2026-{m:02d}": f"2026-{m:02d}" for m in range(1, 13)
}


def guardar_cv_staging(periodo: str, sociedad: str, monto: float) -> dict:
    """
    Inserta o actualiza un registro en staging.cv_real_manual.
    Si ya existe el periodo+sociedad, lo reemplaza (DELETE + INSERT).
    """
    logs = []
    engine = get_engine()
    # Normalizar nombre de sociedad para consistencia con fact_real
    sociedad = sociedad.replace("ACUNA", "ACUÑA") if sociedad == "ACUNA" else sociedad
    try:
        fecha = pd.to_datetime(f"{periodo}-01")
        with engine.begin() as conn:
            r = conn.execute(text("""
                DELETE FROM staging.cv_real_manual
                WHERE TO_CHAR(periodo, 'YYYY-MM') = :p AND sociedad = :s
            """), {"p": periodo, "s": sociedad})
            _log(logs, f"Registros anteriores eliminados: {r.rowcount}")

            conn.execute(text("""
                INSERT INTO staging.cv_real_manual (periodo, sociedad, monto)
                VALUES (:fecha, :s, :m)
            """), {"fecha": fecha, "s": sociedad, "m": monto})
            _log(logs, f"✓ Guardado: {sociedad} {periodo} → ${monto:,.0f}")

        return {"ok": True, "logs": logs, "error": None}
    except Exception as e:
        _log(logs, f"✗ Error: {e}")
        return {"ok": False, "logs": logs, "error": str(e)}


def eliminar_cv_staging(periodo: str, sociedad: str) -> dict:
    """Elimina un registro de staging.cv_real_manual."""
    logs = []
    engine = get_engine()
    sociedad = sociedad.replace("ACUNA", "ACUÑA") if sociedad == "ACUNA" else sociedad
    try:
        with engine.begin() as conn:
            r = conn.execute(text("""
                DELETE FROM staging.cv_real_manual
                WHERE TO_CHAR(periodo, 'YYYY-MM') = :p AND sociedad = :s
            """), {"p": periodo, "s": sociedad})
        _log(logs, f"Eliminado: {sociedad} {periodo} ({r.rowcount} fila)")
        return {"ok": True, "logs": logs, "error": None}
    except Exception as e:
        _log(logs, f"✗ Error: {e}")
        return {"ok": False, "logs": logs, "error": str(e)}


def run_etl_cv_sync() -> dict:
    """
    Sincroniza staging.cv_real_manual → marts.fact_real (COSTO_VAR).
    Para cada periodo+sociedad en staging:
      - Elimina filas con codigo_cuenta = '3.1.01.001' en fact_real
      - Inserta el nuevo monto con fuente = 'CV_MANUAL'
    """
    logs = []
    engine = get_engine()
    try:
        # Migrar registros históricos con 'ACUNA' (sin tilde) a 'ACUÑA' en staging
        with engine.begin() as conn:
            n_migrados = conn.execute(text("""
                UPDATE staging.cv_real_manual
                SET sociedad = 'ACUÑA'
                WHERE sociedad = 'ACUNA'
            """)).rowcount
            if n_migrados:
                _log(logs, f"Staging: {n_migrados} registros 'ACUNA' migrados a 'ACUÑA'")

        # Leer staging
        with engine.connect() as conn:
            df = pd.read_sql(text("""
                SELECT
                    periodo::date                  AS fecha,
                    TO_CHAR(periodo, 'YYYY-MM')    AS periodo_str,
                    sociedad,
                    monto                          AS valor
                FROM staging.cv_real_manual
                WHERE monto > 0
                ORDER BY periodo, sociedad
            """), conn)

        if df.empty:
            _log(logs, "staging.cv_real_manual no tiene registros con monto > 0.")
            return {"ok": True, "n_registros": 0, "logs": logs, "error": None}

        # Normalizar nombre de sociedad: 'ACUNA' sin tilde → 'ACUÑA' con tilde (seguridad extra)
        df["sociedad"] = df["sociedad"].replace({"ACUNA": "ACUÑA"})

        _log(logs, f"Registros en staging: {len(df)}")
        for _, row in df.iterrows():
            _log(logs, f"  {row['periodo_str']} | {row['sociedad']:12s} | ${row['valor']:>15,.0f}")

        # Preparar DataFrame para fact_real
        df_insert = pd.DataFrame({
            "fecha":          df["fecha"],
            "codigo_cuenta":  CODIGO_CUENTA_CV,
            "codigo_cc":      CODIGO_CC_CV,
            "valor":          df["valor"],
            "periodo":        df["periodo_str"],
            "fuente":         "CV_MANUAL",
            "archivo_origen": "staging.cv_real_manual",
            "sociedad":       df["sociedad"],
        })
        df_insert["fecha_id"] = df["fecha"].apply(lambda d: int(d.strftime("%Y%m%d")))

        # Sincronizar período a período
        with engine.begin() as conn:
            for (periodo, sociedad), grupo in df_insert.groupby(["periodo", "sociedad"]):
                # Limpiar tanto 'ACUÑA' como 'ACUNA' (variante sin tilde de datos históricos)
                sociedad_alt = "ACUNA" if sociedad == "ACUÑA" else sociedad
                r = conn.execute(text("""
                    DELETE FROM marts.fact_real
                    WHERE periodo = :p
                      AND sociedad IN (:s, :s_alt)
                      AND codigo_cuenta = :cc
                """), {"p": periodo, "s": sociedad, "s_alt": sociedad_alt, "cc": CODIGO_CUENTA_CV})
                _log(logs, f"  Eliminados {r.rowcount} registros anteriores — {sociedad} {periodo}")

                grupo.to_sql("fact_real", con=conn, schema="marts",
                             if_exists="append", index=False)
                _log(logs, f"  Insertado {len(grupo)} registro — {sociedad} {periodo}  ${grupo['valor'].sum():,.0f}")

        _log(logs, f"✓ Sincronización completa — {len(df_insert)} registros en fact_real")
        _registrar_auditoria(engine, "marts.fact_real", "MULTI", len(df_insert),
                             "ETL CV_REAL SYNC via webapp")

        return {"ok": True, "n_registros": len(df_insert), "logs": logs, "error": None}

    except Exception as e:
        _log(logs, f"✗ Error: {e}")
        return {"ok": False, "n_registros": 0, "logs": logs, "error": str(e)}


# ═══════════════════════════════════════════════════════════════
# PRESUPUESTO DETALLE (item/persona, cuenta, CC)
# staging.ppto_detalle  →  marts.fact_presupuesto (reagregado, CC <> CC-00)
# ═══════════════════════════════════════════════════════════════

_MESES_DET = ["ene", "feb", "mar", "abr", "may", "jun",
              "jul", "ago", "sep", "oct", "nov", "dic"]
FUENTE_PPTO_DET = "PPTO_DETALLE"


def _norm_txt(s) -> str:
    """Normaliza un encabezado: minúsculas, sin acentos, sin espacios extra."""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return s


# Mapa de encabezado normalizado → campo interno
_MAPA_COLS_DET = {
    "sociedad": "sociedad",
    "codigo cc": "codigo_cc",
    "centro de costo": "_centro",
    "codigo cuenta": "codigo_cuenta",
    "nombre cuenta": "_nombre_cuenta",
    "item / nombre": "item",
    "item": "item",
    "tipo": "tipo",
    "notas": "notas",
    **{m: m for m in _MESES_DET},
}


def sincronizar_ppto_detalle(anio: str | int | None = None) -> dict:
    """
    Reagrega staging.ppto_detalle → marts.fact_presupuesto (solo CC <> 'CC-00').
    Suma por (codigo_cuenta, codigo_cc, periodo) — ignora sociedad para mantener
    compatibilidad con la estructura actual de fact_presupuesto. Preserva CC-00
    (Ventas y Costo Variable presupuestados).
    """
    if anio is None:
        anio = date.today().year
    anio = str(anio)
    logs = []
    engine = get_engine()
    try:
        with engine.connect() as conn:
            df = pd.read_sql(text("""
                SELECT codigo_cc, codigo_cuenta,
                       ene, feb, mar, abr, may, jun, jul, ago, sep, oct, nov, dic
                FROM staging.ppto_detalle
                WHERE ano = :a AND codigo_cc <> 'CC-00'
            """), conn, params={"a": int(anio)})

        if df.empty:
            _log(logs, "staging.ppto_detalle sin registros (CC<>CC-00) para el año.")
            # Igual limpiamos lo reagregado previo para no dejar datos colgando
            with engine.begin() as conn:
                conn.execute(text("""
                    DELETE FROM marts.fact_presupuesto
                    WHERE periodo LIKE :a AND codigo_cc <> 'CC-00' AND fuente = :f
                """), {"a": f"{anio}-%", "f": FUENTE_PPTO_DET})
            return {"ok": True, "n_registros": 0, "logs": logs, "error": None}

        # Melt a formato largo (periodo, valor) y agregar
        registros = []
        for _, row in df.iterrows():
            for i, mes in enumerate(_MESES_DET, start=1):
                valor = float(row[mes] or 0)
                if valor == 0:
                    continue
                periodo = f"{anio}-{i:02d}"
                registros.append({
                    "fecha":          pd.to_datetime(f"{periodo}-01"),
                    "codigo_cuenta":  str(row["codigo_cuenta"]).strip(),
                    "codigo_cc":      str(row["codigo_cc"]).strip(),
                    "valor":          valor,
                    "periodo":        periodo,
                    "fuente":         FUENTE_PPTO_DET,
                    "archivo_origen": "staging.ppto_detalle",
                })

        df_long = pd.DataFrame(registros)
        if df_long.empty:
            _log(logs, "Sin montos > 0 para reagregar.")
            return {"ok": True, "n_registros": 0, "logs": logs, "error": None}

        # Agregar por cuenta + CC + periodo
        df_agg = (df_long
                  .groupby(["fecha", "codigo_cuenta", "codigo_cc", "periodo",
                            "fuente", "archivo_origen"], as_index=False)["valor"].sum())
        df_agg["fecha_id"] = df_agg["fecha"].apply(lambda d: int(d.strftime("%Y%m%d")))

        with engine.begin() as conn:
            r = conn.execute(text("""
                DELETE FROM marts.fact_presupuesto
                WHERE periodo LIKE :a AND codigo_cc <> 'CC-00'
            """), {"a": f"{anio}-%"})
            _log(logs, f"fact_presupuesto: {r.rowcount} filas CC<>CC-00 eliminadas")

            df_agg.to_sql("fact_presupuesto", con=conn, schema="marts",
                          if_exists="append", index=False, method="multi")

        _log(logs, f"✓ Reagregadas {len(df_agg)} filas a fact_presupuesto "
                   f"(${df_agg['valor'].sum():,.0f})")
        _registrar_auditoria(engine, "marts.fact_presupuesto", f"{anio}-DET",
                             len(df_agg), "Reagregación ppto_detalle")
        return {"ok": True, "n_registros": len(df_agg), "logs": logs, "error": None}

    except Exception as e:
        _log(logs, f"✗ Error: {e}")
        return {"ok": False, "n_registros": 0, "logs": logs, "error": str(e)}


def run_etl_ppto_detalle(file_bytes: bytes, anio: str | int | None = None) -> dict:
    """
    Carga el Excel plano (hoja PPTO_DETALLE) → staging.ppto_detalle (reemplaza el año)
    y reagrega a marts.fact_presupuesto. Formato esperado de columnas:
    Sociedad, Código CC, Centro de Costo, Código Cuenta, Nombre Cuenta,
    Item / Nombre, Tipo, Notas, Ene..Dic, Total Anual.
    """
    if anio is None:
        anio = date.today().year
    anio = str(anio)
    logs = []
    engine = get_engine()
    try:
        # Elegir hoja PPTO_DETALLE si existe; si no, la primera
        xls = pd.ExcelFile(io.BytesIO(file_bytes))
        hoja = next((h for h in xls.sheet_names if _norm_txt(h) == "ppto detalle"
                     or "detalle" in _norm_txt(h)), xls.sheet_names[0])
        df = pd.read_excel(xls, sheet_name=hoja)
        _log(logs, f"Hoja leída: {hoja} ({len(df)} filas)")

        # Mapear columnas por nombre normalizado
        ren = {}
        for col in df.columns:
            campo = _MAPA_COLS_DET.get(_norm_txt(col))
            if campo:
                ren[col] = campo
        df = df.rename(columns=ren)

        req = ["codigo_cc", "codigo_cuenta"] + _MESES_DET
        faltan = [c for c in req if c not in df.columns]
        if faltan:
            raise ValueError(f"Faltan columnas en el Excel: {faltan}. "
                             f"Revisa que la hoja tenga el formato PPTO_DETALLE.")

        # Defaults para columnas opcionales
        for opt in ["sociedad", "item", "tipo", "notas"]:
            if opt not in df.columns:
                df[opt] = ""

        # Limpieza
        df["codigo_cuenta"] = df["codigo_cuenta"].astype(str).str.strip()
        df["codigo_cc"]     = df["codigo_cc"].astype(str).str.strip()
        df["sociedad"]      = (df["sociedad"].astype(str).str.strip()
                               .replace({"ACUNA": "ACUÑA", "": "Consolidado", "nan": "Consolidado"}))
        for c in ["item", "tipo", "notas"]:
            df[c] = df[c].astype(str).str.strip().replace({"nan": ""})
        for m in _MESES_DET:
            df[m] = pd.to_numeric(df[m], errors="coerce").fillna(0.0)

        # Filtrar filas válidas (cuenta con patrón y CC presente)
        patron = re.compile(r"^\d+\.\d+\.\d+\.\d+$")
        df = df[df["codigo_cuenta"].apply(lambda x: bool(patron.match(x))) &
                df["codigo_cc"].str.startswith("CC-")].copy()
        if df.empty:
            raise ValueError("No se encontraron filas válidas (código de cuenta + CC).")

        df["ano"] = int(anio)
        cols_final = ["ano", "sociedad", "codigo_cc", "codigo_cuenta",
                      "item", "tipo", "notas"] + _MESES_DET
        df_insert = df[cols_final].copy()
        _log(logs, f"Filas válidas a cargar: {len(df_insert)}")

        with engine.begin() as conn:
            r = conn.execute(text("DELETE FROM staging.ppto_detalle WHERE ano = :a"),
                             {"a": int(anio)})
            _log(logs, f"staging.ppto_detalle: {r.rowcount} filas anteriores eliminadas")
            df_insert.to_sql("ppto_detalle", con=conn, schema="staging",
                             if_exists="append", index=False, method="multi")

        _log(logs, f"✓ {len(df_insert)} items cargados en staging.ppto_detalle")

        # Reagregar a fact_presupuesto
        res_sync = sincronizar_ppto_detalle(anio)
        logs.extend(res_sync["logs"])
        if not res_sync["ok"]:
            return {"ok": False, "periodo": f"{anio} (detalle)", "n_registros": len(df_insert),
                    "logs": logs, "error": res_sync["error"]}

        _registrar_auditoria(engine, "staging.ppto_detalle", f"{anio}-DET",
                             len(df_insert), "Carga ppto_detalle via webapp")
        return {"ok": True, "periodo": f"{anio} (detalle)", "n_registros": len(df_insert),
                "logs": logs, "error": None}

    except Exception as e:
        _log(logs, f"✗ Error: {e}")
        return {"ok": False, "periodo": None, "n_registros": 0, "logs": logs, "error": str(e)}


def reemplazar_ppto_detalle(anio: str | int, sociedad: str,
                            df_rows: pd.DataFrame) -> dict:
    """
    Reemplaza los items de una (sociedad, año) en staging.ppto_detalle con las filas
    editadas en la web y reagrega a fact_presupuesto. df_rows debe traer las columnas:
    codigo_cc, codigo_cuenta, item, tipo, notas, ene..dic.
    """
    anio = str(anio)
    logs = []
    engine = get_engine()
    try:
        df = df_rows.copy()
        # Normalizar / validar
        df["codigo_cuenta"] = df["codigo_cuenta"].astype(str).str.strip()
        df["codigo_cc"]     = df["codigo_cc"].astype(str).str.strip()
        patron = re.compile(r"^\d+\.\d+\.\d+\.\d+$")
        df = df[df["codigo_cuenta"].apply(lambda x: bool(patron.match(x))) &
                df["codigo_cc"].str.startswith("CC-")].copy()
        for c in ["item", "tipo", "notas"]:
            if c not in df.columns:
                df[c] = ""
            df[c] = df[c].fillna("").astype(str).str.strip()
        for m in _MESES_DET:
            df[m] = pd.to_numeric(df.get(m, 0), errors="coerce").fillna(0.0)

        df["ano"] = int(anio)
        df["sociedad"] = sociedad
        cols_final = ["ano", "sociedad", "codigo_cc", "codigo_cuenta",
                      "item", "tipo", "notas"] + _MESES_DET
        df_insert = df[cols_final]

        with engine.begin() as conn:
            r = conn.execute(text("""
                DELETE FROM staging.ppto_detalle WHERE ano = :a AND sociedad = :s
            """), {"a": int(anio), "s": sociedad})
            _log(logs, f"Eliminadas {r.rowcount} filas previas — {sociedad} {anio}")
            if not df_insert.empty:
                df_insert.to_sql("ppto_detalle", con=conn, schema="staging",
                                 if_exists="append", index=False, method="multi")
        _log(logs, f"Guardadas {len(df_insert)} filas — {sociedad} {anio}")

        res_sync = sincronizar_ppto_detalle(anio)
        logs.extend(res_sync["logs"])
        return {"ok": res_sync["ok"], "n_registros": len(df_insert),
                "logs": logs, "error": res_sync.get("error")}

    except Exception as e:
        _log(logs, f"✗ Error: {e}")
        return {"ok": False, "n_registros": 0, "logs": logs, "error": str(e)}
